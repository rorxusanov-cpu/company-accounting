from flask import Flask, render_template, request, redirect, url_for, session, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import io
import sqlite3
from datetime import datetime, timedelta
import requests
import os
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)

# ================= CONFIG =================
app.secret_key = os.environ.get("SECRET_KEY", "o'zgartiring-bu-qiymatni")
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "8588085417:AAG1_uFr9irp7-E2fGd20jg0BbxxUopSsH4")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "5703562662")


# ================= TELEGRAM =================
def send_telegram_message(text):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram token yoki chat_id yo'q!")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": text,
        "parse_mode": "HTML"
    }
    try:
        response = requests.post(url, data=payload, timeout=5)
        if not response.json().get("ok"):
            print("Telegram xato javobi:", response.text)
    except Exception as e:
        print("Telegram xato:", e)


def tg_expense(company_name, director, amount, description, created_at, balance=None):
    bal_line = f"\n💳 <b>Qolgan balans:</b> <code>{balance:,} so'm</code>" if balance is not None else ""
    send_telegram_message(
        f"╔══════════════════════╗\n"
        f"║  💸  <b>YANGI XARAJAT</b>  💸  ║\n"
        f"╚══════════════════════╝\n\n"
        f"🏢 <b>Kompaniya:</b> {company_name}\n"
        f"👔 <b>Direktor:</b> {director}\n"
        f"💰 <b>Summa:</b> <code>{amount:,} so'm</code>\n"
        f"📝 <b>Izoh:</b> {description}"
        f"{bal_line}\n"
        f"🕒 <b>Sana:</b> {created_at}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )


def tg_salary(company_name, director, worker_name, amount, note, created_at, balance=None):
    bal_line = f"\n💳 <b>Qolgan balans:</b> <code>{balance:,} so'm</code>" if balance is not None else ""
    send_telegram_message(
        f"╔══════════════════════╗\n"
        f"║  💼  <b>OYLIK MAOSH</b>  💼  ║\n"
        f"╚══════════════════════╝\n\n"
        f"🏢 <b>Kompaniya:</b> {company_name}\n"
        f"👔 <b>Direktor:</b> {director}\n"
        f"👷 <b>Ishchi:</b> {worker_name}\n"
        f"💰 <b>Summa:</b> <code>{amount:,} so'm</code>\n"
        f"📝 <b>Izoh:</b> {note or '—'}"
        f"{bal_line}\n"
        f"🕒 <b>Sana:</b> {created_at}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )


def tg_advance(company_name, director, worker_name, amount, note, created_at, balance=None):
    bal_line = f"\n💳 <b>Qolgan balans:</b> <code>{balance:,} so'm</code>" if balance is not None else ""
    send_telegram_message(
        f"╔══════════════════════╗\n"
        f"║  🏧  <b>AVANS BERILDI</b>  🏧  ║\n"
        f"╚══════════════════════╝\n\n"
        f"🏢 <b>Kompaniya:</b> {company_name}\n"
        f"👔 <b>Direktor:</b> {director}\n"
        f"👷 <b>Ishchi:</b> {worker_name}\n"
        f"💰 <b>Summa:</b> <code>{amount:,} so'm</code>\n"
        f"📝 <b>Izoh:</b> {note or '—'}"
        f"{bal_line}\n"
        f"🕒 <b>Sana:</b> {created_at}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )


# ================= DATABASE =================
def get_db():
    # Render da /data papkasi persistent disk, localda oddiy fayl
    db_path = "/data/users.db" if os.path.exists("/data") else "users.db"
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        role TEXT,
        company_id INTEGER,
        created_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        balance INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)

    # user_id ustuni qo'shildi (avvalgi versiyada yo'q edi — bug)
    c.execute("""
    CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        user_id INTEGER,
        amount INTEGER,
        description TEXT,
        created_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        message TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS workers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_name TEXT,
        position TEXT,
        phone TEXT,
        company_id INTEGER,
        status TEXT DEFAULT 'active',
        monthly_salary INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS salaries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        worker_id INTEGER,
        company_id INTEGER,
        amount INTEGER,
        note TEXT,
        created_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        worker_id INTEGER,
        company_id INTEGER,
        date TEXT,
        status TEXT DEFAULT 'present',
        note TEXT,
        created_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS advances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        worker_id INTEGER,
        company_id INTEGER,
        amount INTEGER,
        note TEXT,
        created_at TEXT
    )
    """)

    # Gullar jadvali
    c.execute("""
    CREATE TABLE IF NOT EXISTS flowers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        name TEXT,
        created_at TEXT
    )
    """)

    # Gul o'lchamlari (razmerlar)
    c.execute("""
    CREATE TABLE IF NOT EXISTS flower_sizes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        flower_id INTEGER,
        size_name TEXT,
        created_at TEXT
    )
    """)

    # Chiqqan gullar (harvests)
    c.execute("""
    CREATE TABLE IF NOT EXISTS flower_harvests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        flower_id INTEGER,
        size_id INTEGER,
        company_id INTEGER,
        quantity INTEGER,
        note TEXT,
        created_at TEXT
    )
    """)

    # monthly_salary ustunini mavjud workers jadvaliga qo'shish (migration)
    try:
        c.execute("ALTER TABLE workers ADD COLUMN status TEXT DEFAULT 'active'")
    except:
        pass
    try:
        c.execute("ALTER TABLE workers ADD COLUMN monthly_salary INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE expenses ADD COLUMN is_edited INTEGER DEFAULT 0")
    except:
        pass
    try:
        c.execute("ALTER TABLE expenses ADD COLUMN edited_at TEXT")
    except:
        pass
    try:
        c.execute("ALTER TABLE expenses ADD COLUMN original_amount INTEGER")
    except:
        pass
    try:
        c.execute("ALTER TABLE expenses ADD COLUMN original_description TEXT")
    except:
        pass

    # Default admin: parol hash qilingan holda saqlanadi
    c.execute("SELECT * FROM users WHERE role='admin'")
    if not c.fetchone():
        c.execute("""
            INSERT INTO users (username, password, role, created_at)
            VALUES ('admin', ?, 'admin', ?)
        """, (
            generate_password_hash("Admin@12345"),  # Kuchli default parol
            datetime.now().strftime("%Y-%m-%d %H:%M")
        ))

    conn.commit()
    conn.close()


# ================= AUTH =================
@app.route("/", methods=["GET", "POST"])
def login():
    error = ""

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        conn = get_db()
        c = conn.cursor()
        c.execute(
            "SELECT id, username, password, role, company_id FROM users WHERE username=?",
            (username,)
        )
        user = c.fetchone()
        conn.close()

        # check_password_hash — parolni xavfsiz tekshiradi
        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["user"] = user["username"]
            session["role"] = user["role"]
            session["company_id"] = user["company_id"]
            return redirect(url_for("dashboard"))
        else:
            error = "Login yoki parol noto'g'ri!"

    return render_template("login.html", error=error)


@app.route("/signup", methods=["GET", "POST"])
def signup():
    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if not username or not password:
            error = "Username va parol bo'sh bo'lishi mumkin emas!"
            return render_template("signup.html", error=error)

        if len(password) < 6:
            error = "Parol kamida 6 ta belgidan iborat bo'lishi kerak!"
            return render_template("signup.html", error=error)

        try:
            conn = get_db()
            c = conn.cursor()
            c.execute("""
                INSERT INTO users (username, password, role, created_at)
                VALUES (?, ?, 'director', ?)
            """, (
                username,
                generate_password_hash(password),  # Parol hash qilinadi
                datetime.now().strftime("%Y-%m-%d %H:%M")
            ))
            conn.commit()
            conn.close()
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            error = "Bu login mavjud!"
        except Exception as e:
            print("Signup xato:", e)
            error = "Xatolik yuz berdi, qayta urinib ko'ring."

    return render_template("signup.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ================= DASHBOARD =================
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()

    role = session.get("role")
    username = session.get("user")

    period = request.args.get("period")
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    total_balance = 0
    total_expenses = 0
    profit = 0
    companies_count = 0
    labels = []
    values = []
    warning = None

    # Qo'shimcha statistika o'zgaruvchilari
    top_companies = []
    recent_expenses = []
    workers_count = 0
    month_expenses = 0
    today_expenses = 0
    directors_count = 0

    # ================= ADMIN =================
    if role == "admin":
        c.execute("SELECT IFNULL(SUM(balance),0) FROM companies")
        total_balance = c.fetchone()[0]

        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses")
        total_expenses = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM companies")
        companies_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM users WHERE role='director'")
        directors_count = c.fetchone()[0]

        c.execute("SELECT COUNT(*) FROM workers")
        workers_count = c.fetchone()[0]

        # Bu oylik xarajat
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m', created_at)=strftime('%Y-%m','now')")
        month_expenses = c.fetchone()[0]

        # Bugungi xarajat
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE date(created_at)=date('now')")
        today_expenses = c.fetchone()[0]

        # Top 5 kompaniya (eng ko'p xarajat)
        c.execute("""
            SELECT c.name, c.balance, IFNULL(SUM(e.amount),0) as exp
            FROM companies c
            LEFT JOIN expenses e ON e.company_id=c.id
            GROUP BY c.id ORDER BY exp DESC LIMIT 5
        """)
        top_companies = c.fetchall()

        # So'nggi 7 ta xarajat
        c.execute("""
            SELECT e.amount, e.description, e.created_at, c.name, u.username
            FROM expenses e
            JOIN companies c ON e.company_id=c.id
            LEFT JOIN users u ON e.user_id=u.id
            ORDER BY e.created_at DESC LIMIT 7
        """)
        recent_expenses = c.fetchall()

        profit = total_balance

        if period == "day":
            c.execute("""
                SELECT strftime('%H', created_at), SUM(amount)
                FROM expenses
                WHERE date(created_at)=date('now')
                GROUP BY strftime('%H', created_at)
            """)
            data = c.fetchall()
            labels = [f"{row[0]}:00" for row in data]
            values = [row[1] for row in data]

        elif period == "month":
            c.execute("""
                SELECT date(created_at), SUM(amount)
                FROM expenses
                WHERE strftime('%Y-%m', created_at)=strftime('%Y-%m','now')
                GROUP BY date(created_at)
            """)
            data = c.fetchall()
            labels = [row[0] for row in data]
            values = [row[1] for row in data]

        elif period == "year":
            c.execute("""
                SELECT strftime('%Y-%m', created_at), SUM(amount)
                FROM expenses
                WHERE strftime('%Y', created_at)=strftime('%Y','now')
                GROUP BY strftime('%Y-%m', created_at)
            """)
            data = c.fetchall()
            labels = [row[0] for row in data]
            values = [row[1] for row in data]

        elif date_from and date_to:
            c.execute("""
                SELECT date(created_at), SUM(amount)
                FROM expenses
                WHERE date(created_at) BETWEEN ? AND ?
                GROUP BY date(created_at)
            """, (date_from, date_to))
            data = c.fetchall()
            labels = [row[0] for row in data]
            values = [row[1] for row in data]

        else:
            for i in range(6, -1, -1):
                day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                c.execute(
                    "SELECT IFNULL(SUM(amount),0) FROM expenses WHERE date(created_at)=?",
                    (day,)
                )
                labels.append(day)
                values.append(c.fetchone()[0])

    # ================= DIRECTOR =================
    else:
        c.execute(
            "SELECT company_id FROM users WHERE username=?",
            (username,)
        )
        row = c.fetchone()

        if not row or not row["company_id"]:
            warning = "Siz kompaniyaga biriktirilmagansiz"
        else:
            company_id = row["company_id"]

            c.execute(
                "SELECT IFNULL(balance,0) FROM companies WHERE id=?",
                (company_id,)
            )
            total_balance = c.fetchone()[0]

            c.execute(
                "SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=?",
                (company_id,)
            )
            total_expenses = c.fetchone()[0]

            profit = total_balance - total_expenses
            companies_count = 1

            if period == "day":
                c.execute("""
                    SELECT strftime('%H', created_at), SUM(amount)
                    FROM expenses
                    WHERE company_id=? AND date(created_at)=date('now')
                    GROUP BY strftime('%H', created_at)
                """, (company_id,))
                data = c.fetchall()
                labels = [f"{row[0]}:00" for row in data]
                values = [row[1] for row in data]

            elif period == "month":
                c.execute("""
                    SELECT date(created_at), SUM(amount)
                    FROM expenses
                    WHERE company_id=? AND strftime('%Y-%m', created_at)=strftime('%Y-%m','now')
                    GROUP BY date(created_at)
                """, (company_id,))
                data = c.fetchall()
                labels = [row[0] for row in data]
                values = [row[1] for row in data]

            elif period == "year":
                c.execute("""
                    SELECT strftime('%Y-%m', created_at), SUM(amount)
                    FROM expenses
                    WHERE company_id=? AND strftime('%Y', created_at)=strftime('%Y','now')
                    GROUP BY strftime('%Y-%m', created_at)
                """, (company_id,))
                data = c.fetchall()
                labels = [row[0] for row in data]
                values = [row[1] for row in data]

            elif date_from and date_to:
                c.execute("""
                    SELECT date(created_at), SUM(amount)
                    FROM expenses
                    WHERE company_id=? AND date(created_at) BETWEEN ? AND ?
                    GROUP BY date(created_at)
                """, (company_id, date_from, date_to))
                data = c.fetchall()
                labels = [row[0] for row in data]
                values = [row[1] for row in data]

            else:
                for i in range(6, -1, -1):
                    day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                    c.execute(
                        """SELECT IFNULL(SUM(amount),0)
                           FROM expenses
                           WHERE company_id=? AND date(created_at)=?""",
                        (company_id, day)
                    )
                    labels.append(day)
                    values.append(c.fetchone()[0])

    conn.close()

    return render_template(
        "dashboard.html",
        total_balance=total_balance,
        total_expenses=total_expenses,
        profit=profit,
        companies_count=companies_count,
        directors_count=directors_count,
        workers_count=workers_count,
        month_expenses=month_expenses,
        today_expenses=today_expenses,
        top_companies=top_companies,
        recent_expenses=recent_expenses,
        labels=labels,
        values=values,
        warning=warning
    )


# ================= ADMIN : COMPANIES =================
@app.route("/admin/companies", methods=["GET", "POST"])
def admin_companies():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        balance = int(request.form.get("balance", 0))

        if name:
            c.execute(
                "INSERT INTO companies (name, balance, created_at) VALUES (?, ?, ?)",
                (name, balance, datetime.now().strftime("%Y-%m-%d %H:%M"))
            )
            conn.commit()

        conn.close()
        return redirect(url_for("admin_companies"))

    c.execute("""
        SELECT
            c.id,
            c.name,
            c.balance,
            u.username AS director,
            IFNULL(SUM(e.amount), 0) AS total_expense
        FROM companies c
        LEFT JOIN users u ON u.company_id = c.id AND u.role = 'director'
        LEFT JOIN expenses e ON e.company_id = c.id
        GROUP BY c.id
        ORDER BY c.id DESC
    """)
    companies = c.fetchall()
    conn.close()

    return render_template("admin_companies.html", companies=companies)


# ================= EXPENSES (DIRECTOR) =================
@app.route("/expenses", methods=["GET", "POST"])
def expenses():
    if "user" not in session:
        return redirect(url_for("login"))

    if session.get("role") != "director":
        return redirect(url_for("dashboard"))

    conn = get_db()
    c = conn.cursor()

    c.execute(
        "SELECT id, company_id FROM users WHERE username=?",
        (session["user"],)
    )
    user_row = c.fetchone()

    if not user_row or not user_row["company_id"]:
        conn.close()
        return render_template(
            "expenses.html",
            expenses=[],
            balance=0,
            error="Siz kompaniyaga biriktirilmagansiz ❌"
        )

    user_id = user_row["id"]
    company_id = user_row["company_id"]

    if request.method == "POST":
        try:
            amount = int(request.form.get("amount", 0))
        except ValueError:
            amount = 0

        description = request.form.get("description", "").strip()
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")

        c.execute(
            "SELECT balance, name FROM companies WHERE id=?",
            (company_id,)
        )
        company_row = c.fetchone()
        balance = company_row["balance"]
        company_name = company_row["name"]

        if amount <= 0:
            conn.close()
            return render_template(
                "expenses.html",
                expenses=[],
                balance=balance,
                error="Xarajat summasi noto'g'ri ❌"
            )

        if amount > balance:
            conn.close()
            return render_template(
                "expenses.html",
                expenses=[],
                balance=balance,
                error="Balans yetarli emas ❌"
            )

        # user_id ham yoziladi (tuzatildi)
        c.execute("""
            INSERT INTO expenses (company_id, user_id, amount, description, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (company_id, user_id, amount, description, created_at))

        c.execute(
            "UPDATE companies SET balance = balance - ? WHERE id=?",
            (amount, company_id)
        )

        conn.commit()

        # Yangilangan balansni olamiz
        c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
        new_bal = c.fetchone()
        tg_expense(company_name, session['user'], amount, description, created_at,
                   balance=new_bal["balance"] if new_bal else None)

        conn.close()
        return redirect(url_for("expenses"))

    c.execute("""
        SELECT id, amount, description, created_at, is_edited, edited_at, original_amount, original_description
        FROM expenses
        WHERE company_id=?
        ORDER BY id DESC
    """, (company_id,))
    expense_list = c.fetchall()

    c.execute(
        "SELECT balance FROM companies WHERE id=?",
        (company_id,)
    )
    balance = c.fetchone()["balance"]

    conn.close()

    return render_template(
        "expenses.html",
        expenses=expense_list,
        balance=balance,
        error=None
    )


# ================= XARAJAT TAHRIRLASH (DIRECTOR) =================
@app.route("/expenses/edit/<int:expense_id>", methods=["GET", "POST"])
def edit_expense(expense_id):
    if "user" not in session or session.get("role") != "director":
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()

    # Direktorning kompaniyasini DB dan olamiz
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else None

    c.execute("SELECT * FROM expenses WHERE id=? AND company_id=?", (expense_id, company_id))
    expense = c.fetchone()
    if not expense:
        conn.close()
        return "Xarajat topilmadi", 404

    if request.method == "POST":
        new_amount = int(request.form.get("amount", expense["amount"]) or expense["amount"])
        new_desc = request.form.get("description", expense["description"]).strip()
        edited_at = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Farqni balansga qaytaramiz yoki yechamiz
        diff = new_amount - expense["amount"]
        c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
        bal = c.fetchone()

        if diff > 0 and bal["balance"] < diff:
            conn.close()
            return redirect(url_for("expenses"))

        # Asl qiymatlarni saqlaymiz (agar birinchi tahrirlash bo'lsa)
        orig_amount = expense["original_amount"] if expense["is_edited"] else expense["amount"]
        orig_desc = expense["original_description"] if expense["is_edited"] else expense["description"]

        c.execute("""
            UPDATE expenses SET
                amount=?, description=?, is_edited=1,
                edited_at=?, original_amount=?, original_description=?
            WHERE id=?
        """, (new_amount, new_desc, edited_at, orig_amount, orig_desc, expense_id))

        # Balansni farqqa qarab yangilaymiz
        if diff != 0:
            c.execute("UPDATE companies SET balance = balance - ? WHERE id=?", (diff, company_id))

        conn.commit()

        # Admin ga Telegram xabari
        c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
        comp = c.fetchone()
        c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
        new_bal = c.fetchone()

        send_telegram_message(
            f"╔══════════════════════╗\n"
            f"║  ✏️  <b>XARAJAT TAHRIRLANDI</b>  ✏️  ║\n"
            f"╚══════════════════════╝\n\n"
            f"🏢 <b>Kompaniya:</b> {comp['name'] if comp else '—'}\n"
            f"👔 <b>Direktor:</b> {session['user']}\n"
            f"📋 <b>Asl izoh:</b> {orig_desc}\n"
            f"💰 <b>Asl summa:</b> <code>{orig_amount:,} so'm</code>\n"
            f"📝 <b>Yangi izoh:</b> {new_desc}\n"
            f"💰 <b>Yangi summa:</b> <code>{new_amount:,} so'm</code>\n"
            f"💳 <b>Qolgan balans:</b> <code>{new_bal['balance']:,} so'm</code>\n"
            f"🕒 <b>Sana:</b> {edited_at}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━"
        )

        conn.close()
        return redirect(url_for("expenses"))

    conn.close()
    return render_template("edit_expense.html", expense=expense)


# ================= ADMIN : DIRECTOR DETAIL =================
@app.route("/admin/director/<int:director_id>", methods=["GET", "POST"])
def admin_director_detail(director_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT
            u.id,
            u.username,
            u.created_at,
            c.name AS company,
            c.id AS company_id_val
        FROM users u
        LEFT JOIN companies c ON u.company_id = c.id
        WHERE u.id = ?
    """, (director_id,))
    director = c.fetchone()

    if not director:
        conn.close()
        return "Direktor topilmadi ❌"

    if request.method == "POST":
        company_id = request.form.get("company_id")
        if company_id:
            c.execute(
                "UPDATE users SET company_id=? WHERE id=?",
                (company_id, director_id)
            )
            conn.commit()
        conn.close()
        return redirect(url_for("admin_director_detail", director_id=director_id))

    c.execute("SELECT id, name FROM companies")
    companies = c.fetchall()

    company_balance = 0
    total_expenses = 0
    recent_expenses = []
    chart_labels = []
    chart_values = []

    if director["company_id_val"]:
        c.execute("SELECT balance FROM companies WHERE id=?", (director["company_id_val"],))
        row = c.fetchone()
        if row:
            company_balance = row["balance"]

        c.execute("SELECT IFNULL(SUM(amount), 0) FROM expenses WHERE company_id=?", (director["company_id_val"],))
        total_expenses = c.fetchone()[0]

        # Oxirgi 10 ta xarajat
        c.execute("""
            SELECT amount, description, created_at
            FROM expenses WHERE company_id=?
            ORDER BY id DESC LIMIT 10
        """, (director["company_id_val"],))
        recent_expenses = c.fetchall()

        # Oxirgi 7 kunlik grafik
        for i in range(6, -1, -1):
            day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=? AND date(created_at)=?",
                      (director["company_id_val"], day))
            chart_labels.append(day[-5:])
            chart_values.append(c.fetchone()[0])

    conn.close()

    return render_template(
        "admin_director_detail.html",
        director=director,
        companies=companies,
        company_balance=company_balance,
        total_expenses=total_expenses,
        recent_expenses=recent_expenses,
        chart_labels=chart_labels,
        chart_values=chart_values
    )


# ================= ADMIN : DIRECTORS =================
@app.route("/admin/directors")
def admin_directors():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT
            u.id,
            u.username,
            c.name AS company,
            u.created_at
        FROM users u
        LEFT JOIN companies c ON u.company_id = c.id
        WHERE u.role = 'director'
        ORDER BY u.id DESC
    """)
    directors = c.fetchall()
    conn.close()

    return render_template("admin_directors.html", directors=directors)


# ================= ADMIN : COMPANY DETAIL =================
@app.route("/admin/company/<int:company_id>")
def admin_company_detail(company_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute(
        "SELECT id, name, balance, created_at FROM companies WHERE id=?",
        (company_id,)
    )
    company = c.fetchone()
    if not company:
        conn.close()
        return "Kompaniya topilmadi ❌"

    c.execute(
        "SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=?",
        (company_id,)
    )
    total_expenses = c.fetchone()[0]

    # Tuzatildi: har bir direktor o'zining xarajatlarini ko'rsatadi
    c.execute("""
        SELECT
            u.id,
            u.username,
            IFNULL(SUM(e.amount),0) AS total
        FROM users u
        LEFT JOIN expenses e ON e.user_id = u.id
        WHERE u.company_id=? AND u.role='director'
        GROUP BY u.id
    """, (company_id,))
    directors = c.fetchall()

    period = request.args.get("period")
    from_date = request.args.get("from")
    to_date = request.args.get("to")

    query = """
        SELECT
            u.username,
            e.amount,
            e.description,
            e.created_at
        FROM expenses e
        LEFT JOIN users u ON u.id = e.user_id
        WHERE e.company_id=?
    """
    params = [company_id]

    if period == "day":
        query += " AND date(e.created_at)=date('now')"
    elif period == "month":
        query += " AND strftime('%Y-%m', e.created_at)=strftime('%Y-%m','now')"
    elif from_date and to_date:
        query += " AND date(e.created_at) BETWEEN ? AND ?"
        params.extend([from_date, to_date])

    query += " ORDER BY e.created_at DESC"
    c.execute(query, params)
    expense_list = c.fetchall()

    # Oxirgi 7 kunlik grafik
    chart_labels = []
    chart_values = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        c.execute(
            "SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=? AND date(created_at)=?",
            (company_id, day)
        )
        chart_labels.append(day[-5:])
        chart_values.append(c.fetchone()[0])

    # Oylik xarajatlar (oxirgi 6 oy)
    c.execute("""
        SELECT strftime('%Y-%m', created_at) as month, IFNULL(SUM(amount),0)
        FROM expenses WHERE company_id=?
        GROUP BY month ORDER BY month DESC LIMIT 6
    """, (company_id,))
    monthly = c.fetchall()

    # Balans to'ldirish uchun POST
    conn.close()

    return render_template(
        "admin_company_detail.html",
        company=company,
        total_expenses=total_expenses,
        directors=directors,
        expenses=expense_list,
        chart_labels=chart_labels,
        chart_values=chart_values,
        monthly=monthly
    )


# ================= ADMIN : EXPENSES =================
@app.route("/admin/expenses")
def admin_expenses():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    day = request.args.get("day")
    month = request.args.get("month")
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT
            c.name AS company,
            CAST(e.amount AS INTEGER) AS amount,
            e.description,
            e.created_at
        FROM expenses e
        JOIN companies c ON e.company_id = c.id
        WHERE 1=1
    """
    params = []

    if day:
        query += " AND date(e.created_at) = ?"
        params.append(day)
    elif month:
        query += " AND strftime('%Y-%m', e.created_at) = ?"
        params.append(month)
    elif date_from and date_to:
        query += " AND date(e.created_at) BETWEEN ? AND ?"
        params.extend([date_from, date_to])

    query += " ORDER BY e.id DESC"

    c.execute(query, params)
    expense_list = c.fetchall()
    conn.close()

    return render_template("admin_expenses.html", expenses=expense_list)


# ================= ADMIN : EXPENSES BY COMPANY =================
@app.route("/admin/expenses/companies")
def admin_expenses_by_company():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT
            c.id,
            c.name,
            IFNULL(SUM(e.amount),0) AS total_expense
        FROM companies c
        LEFT JOIN expenses e ON e.company_id = c.id
        GROUP BY c.id, c.name
        ORDER BY total_expense DESC
    """)
    companies = c.fetchall()
    conn.close()

    return render_template("admin_expenses_companies.html", companies=companies)


# ================= ADMIN : ASSIGN DIRECTOR =================
@app.route("/admin/assign-director", methods=["POST"])
def assign_director():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    director_id = request.form.get("director_id")
    company_id = request.form.get("company_id")

    if not director_id or not company_id:
        return "Ma'lumotlar yetarli emas ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        UPDATE users
        SET company_id=?
        WHERE id=? AND role='director'
    """, (company_id, director_id))

    conn.commit()
    conn.close()

    return redirect(url_for("admin_director_detail", director_id=director_id))


# ================= ADMIN : UNASSIGN DIRECTOR =================
@app.route("/admin/unassign-director/<int:director_id>")
def unassign_director(director_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute(
        "UPDATE users SET company_id=NULL WHERE id=?",
        (director_id,)
    )

    conn.commit()
    conn.close()

    return redirect(url_for("admin_director_detail", director_id=director_id))


# ================= ADMIN : REPORTS =================
@app.route("/admin/reports")
def admin_reports():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    # Kompaniyalar bo'yicha jami xarajat
    c.execute("""
        SELECT c.name, IFNULL(SUM(e.amount),0)
        FROM companies c
        LEFT JOIN expenses e ON e.company_id=c.id
        GROUP BY c.name ORDER BY 2 DESC
    """)
    company_data = c.fetchall()

    # Umumiy statistika
    c.execute("SELECT IFNULL(SUM(balance),0) FROM companies")
    total_balance = c.fetchone()[0]

    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses")
    total_expenses = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM companies")
    total_companies = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM users WHERE role='director'")
    total_directors = c.fetchone()[0]

    # Oxirgi 12 oylik trend
    c.execute("""
        SELECT strftime('%Y-%m', created_at) as month, IFNULL(SUM(amount),0)
        FROM expenses
        GROUP BY month ORDER BY month DESC LIMIT 12
    """)
    monthly_raw = c.fetchall()
    monthly_data = list(reversed(monthly_raw))

    # Eng ko'p xarajat qilgan top 5 direktor
    c.execute("""
        SELECT u.username, c.name, IFNULL(SUM(e.amount),0) as total
        FROM users u
        LEFT JOIN expenses e ON e.user_id = u.id
        LEFT JOIN companies c ON u.company_id = c.id
        WHERE u.role='director'
        GROUP BY u.id ORDER BY total DESC LIMIT 5
    """)
    top_directors = c.fetchall()

    # Haftalik xarajat (oxirgi 7 kun)
    weekly_labels = []
    weekly_values = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE date(created_at)=?", (day,))
        weekly_labels.append(day[-5:])
        weekly_values.append(c.fetchone()[0])

    # Kompaniyalar ro'yxati (Telegram yuborish modal uchun)
    c.execute("SELECT id, name FROM companies ORDER BY name")
    companies_list = c.fetchall()

    conn.close()

    return render_template(
        "admin_reports.html",
        labels=[row[0] for row in company_data],
        values=[row[1] for row in company_data],
        total_balance=total_balance,
        total_expenses=total_expenses,
        total_companies=total_companies,
        total_directors=total_directors,
        monthly_labels=[row[0] for row in monthly_data],
        monthly_values=[row[1] for row in monthly_data],
        top_directors=top_directors,
        weekly_labels=weekly_labels,
        weekly_values=weekly_values,
        companies_list=companies_list,
    )


# ================= ADMIN : BALANCES =================
@app.route("/admin/balances", methods=["GET", "POST"])
def admin_balances():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        company_id = request.form.get("company_id")
        try:
            amount = int(request.form.get("amount", 0))
        except ValueError:
            amount = 0

        if company_id and amount > 0:
            c.execute(
                "UPDATE companies SET balance = balance + ? WHERE id=?",
                (amount, company_id)
            )
            conn.commit()

    c.execute("SELECT id, name, balance FROM companies")
    companies = c.fetchall()
    conn.close()

    return render_template("admin_balances.html", companies=companies)


# ================= ADMIN : COMPANY DIRECTORS =================
@app.route("/admin/company/<int:company_id>/directors")
def admin_company_directors(company_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    company = c.fetchone()
    if not company:
        conn.close()
        return "Kompaniya topilmadi ❌"

    # Tuzatildi: e.user_id = u.id orqali to'g'ri join
    c.execute("""
        SELECT
            u.username,
            IFNULL(SUM(e.amount), 0) AS total
        FROM users u
        LEFT JOIN expenses e ON e.user_id = u.id
        WHERE u.company_id = ? AND u.role='director'
        GROUP BY u.username
    """, (company_id,))
    data = c.fetchall()
    conn.close()

    return render_template(
        "admin_company_directors.html",
        company_name=company["name"],
        data=data
    )


# ================= CLEAR NOTIFICATION =================
@app.route("/clear-notification", methods=["POST"])
def clear_notification():
    session.pop("last_expense", None)
    return redirect(request.referrer or url_for("dashboard"))


# ================= CONTEXT PROCESSOR =================
@app.context_processor
def inject_notifications():
    return {
        "notifications": [],
        "last_expense": session.get("last_expense")
    }





# ================= EXCEL EXPORT =================


def _xl_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_cell(cell, bold=False, color="0F172A", bg=None, align="left",
             size=10, number_fmt=None):
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_fmt:
        cell.number_format = number_fmt
    cell.border = _xl_border()

def _xl_header(ws, title, subtitle, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value = title
    c.font = Font(name="Arial", bold=True, size=18, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0A0F1E")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 42
    ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
    c = ws["A3"]
    c.value = subtitle
    c.font = Font(name="Arial", size=10, color="94A3B8")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[3].height = 22
    ws.merge_cells(f"A4:{get_column_letter(ncols)}4")
    ws["A4"].fill = PatternFill("solid", fgColor="0A0F1E")
    ws.row_dimensions[4].height = 6

def _xl_col_headers(ws, row, headers, bg="3B6EF6"):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _xl_border()
    ws.row_dimensions[row].height = 26

def _build_xlsx(rows, sheet_title, report_label):
    """rows: list of (company, director, amount, description, created_at)"""
    from collections import defaultdict
    wb = Workbook()

    # ===== SHEET 1: XARAJATLAR =====
    ws = wb.active
    ws.title = sheet_title
    ncols = 5
    _xl_header(ws, f"💸  {report_label.upper()}",
               f"Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}", ncols)
    widths = [5, 22, 18, 34, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _xl_col_headers(ws, 5, ["#", "Kompaniya", "Summa (so'm)", "Izoh", "Sana"])

    DS = 6
    for idx, row in enumerate(rows, 1):
        r = DS + idx - 1
        bg = "F0F4FF" if idx % 2 == 0 else "FFFFFF"
        for col, val in enumerate([idx, row[0], row[2], row[3], row[4]], 1):
            c = ws.cell(row=r, column=col, value=val)
            _xl_cell(c, bg=bg, align="center" if col == 1 else "left")
            if col == 3:
                _xl_cell(c, bold=True, color="EF4444", bg=bg,
                         align="right", number_fmt="#,##0")
            if col == 5:
                _xl_cell(c, color="64748B", bg=bg)
        ws.row_dimensions[r].height = 20

    DE = DS + len(rows) - 1
    if rows:
        tr = DE + 1
        ws.merge_cells(f"A{tr}:{get_column_letter(2)}{tr}")
        c = ws[f"A{tr}"]
        c.value = "JAMI XARAJAT"
        _xl_cell(c, bold=True, color="FFFFFF", bg="0A0F1E",
                 align="center", size=11)
        tc = ws.cell(row=tr, column=3,
                     value=f"=SUM(C{DS}:C{DE})")
        _xl_cell(tc, bold=True, color="00D4AA", bg="0A0F1E",
                 align="right", size=12, number_fmt="#,##0")
        for col in [4, 5]:
            ws.cell(row=tr, column=col).fill = PatternFill("solid", fgColor="0A0F1E")
            ws.cell(row=tr, column=col).border = _xl_border()
        ws.row_dimensions[tr].height = 28
    ws.freeze_panes = "A6"

    # ===== SHEET 2: KOMPANIYALAR YIG'MA =====
    ws2 = wb.create_sheet("Kompaniyalar")
    company_totals = defaultdict(int)
    for row in rows:
        company_totals[row[0]] += row[2]
    sorted_companies = sorted(company_totals.items(), key=lambda x: -x[1])

    nc2 = 3
    _xl_header(ws2, "🏢  KOMPANIYALAR BO'YICHA YIG'MA",
               f"Sana: {datetime.now().strftime('%d.%m.%Y')}", nc2)
    for i, w in enumerate([5, 26, 20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    _xl_col_headers(ws2, 5, ["#", "Kompaniya", "Jami xarajat"], bg="00B894")

    DS2 = 6
    for idx, (comp, total) in enumerate(sorted_companies, 1):
        r = DS2 + idx - 1
        bg = "F0FDF4" if idx % 2 == 0 else "FFFFFF"
        for col, val in enumerate([idx, comp, total], 1):
            c = ws2.cell(row=r, column=col, value=val)
            _xl_cell(c, bg=bg, align="center" if col == 1 else "left")
            if col == 3:
                _xl_cell(c, bold=True, color="EF4444", bg=bg,
                         align="right", number_fmt="#,##0")
        ws2.row_dimensions[r].height = 20

    DE2 = DS2 + len(sorted_companies) - 1
    if sorted_companies:
        tr2 = DE2 + 1
        ws2[f"A{tr2}"].value = "JAMI"
        _xl_cell(ws2[f"A{tr2}"], bold=True, color="FFFFFF",
                 bg="0A0F1E", align="center", size=11)
        tc2 = ws2.cell(row=tr2, column=3,
                       value=f"=SUM(C{DS2}:C{DE2})")
        _xl_cell(tc2, bold=True, color="00D4AA", bg="0A0F1E",
                 align="right", size=12, number_fmt="#,##0")
        ws2.cell(row=tr2, column=2).fill = PatternFill("solid", fgColor="0A0F1E")
        ws2.cell(row=tr2, column=2).border = _xl_border()
        ws2.row_dimensions[tr2].height = 28
    ws2.freeze_panes = "A6"

    # ===== SHEET 3: GRAFIK =====
    ws3 = wb.create_sheet("Grafik")
    _xl_header(ws3, "📊  GRAFIK TAHLIL",
               "Kompaniyalar bo'yicha xarajatlar taqqoslash", 4)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 20
    _xl_col_headers(ws3, 5, ["Kompaniya", "Xarajat (so'm)"])
    for i, (comp, total) in enumerate(sorted_companies, 6):
        ws3.cell(row=i, column=1, value=comp).font = Font(name="Arial", size=10)
        c = ws3.cell(row=i, column=2, value=total)
        c.number_format = "#,##0"
        c.font = Font(name="Arial", size=10, color="EF4444", bold=True)
        ws3.row_dimensions[i].height = 20

    if sorted_companies:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Kompaniyalar bo'yicha xarajatlar"
        chart.y_axis.title = "So'm"
        chart.style = 10
        chart.width = 22
        chart.height = 14
        max_row = 5 + len(sorted_companies)
        data_ref = Reference(ws3, min_col=2, min_row=5, max_row=max_row)
        cats = Reference(ws3, min_col=1, min_row=6, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        palette = ["3B6EF6","00D4AA","F59E0B","EF4444","8B5CF6","06B6D4"]
        for i in range(len(sorted_companies)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = palette[i % len(palette)]
            chart.series[0].dPt.append(pt)
        ws3.add_chart(chart, f"A{max_row + 3}")

    return wb


@app.route("/admin/export/expenses")
def export_all_expenses():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT c.name, u.username, e.amount, e.description, e.created_at
        FROM expenses e
        JOIN companies c ON e.company_id = c.id
        LEFT JOIN users u ON u.id = e.user_id
        ORDER BY e.created_at DESC
    """)
    rows = c.fetchall()
    conn.close()
    wb = _build_xlsx(rows, "Xarajatlar", "Barcha xarajatlar hisoboti")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"xarajatlar_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/admin/export/company/<int:company_id>")
def export_company_expenses(company_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
    company = c.fetchone()
    if not company:
        conn.close()
        return "Kompaniya topilmadi ❌"
    c.execute("""
        SELECT c.name, u.username, e.amount, e.description, e.created_at
        FROM expenses e
        JOIN companies c ON e.company_id = c.id
        LEFT JOIN users u ON u.id = e.user_id
        WHERE e.company_id = ?
        ORDER BY e.created_at DESC
    """, (company_id,))
    rows = c.fetchall()
    conn.close()
    wb = _build_xlsx(rows, "Xarajatlar", f"{company['name']} xarajatlari")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{company['name']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/director/export/expenses")
def export_director_expenses():
    if "user" not in session or session.get("role") != "director":
        return redirect(url_for("login"))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, company_id FROM users WHERE username=?", (session["user"],))
    user_row = c.fetchone()
    if not user_row or not user_row["company_id"]:
        conn.close()
        return "Kompaniyaga biriktirilmagansiz ❌"
    c.execute("""
        SELECT c.name, u.username, e.amount, e.description, e.created_at
        FROM expenses e
        JOIN companies c ON e.company_id = c.id
        JOIN users u ON u.id = e.user_id
        WHERE e.company_id = ? AND e.user_id = ?
        ORDER BY e.created_at DESC
    """, (user_row["company_id"], user_row["id"]))
    rows = c.fetchall()
    conn.close()
    wb = _build_xlsx(rows, "Xarajatlarim", f"{session['user']} xarajatlari")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{session['user']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ================= ISHCHILAR =================
@app.route("/workers", methods=["GET", "POST"])
def workers():
    if "user" not in session:
        return redirect(url_for("login"))
    if session.get("role") != "director":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    # Direktorning kompaniyasini topamiz
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    row = c.fetchone()
    company_id = row["company_id"] if row else None

    error = ""

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add":
            full_name = request.form.get("full_name", "").strip()
            position = request.form.get("position", "").strip()
            phone = request.form.get("phone", "").strip()
            monthly_salary = int(request.form.get("monthly_salary", 0) or 0)
            if full_name:
                c.execute(
                    "INSERT INTO workers (full_name, position, phone, company_id, monthly_salary, created_at) VALUES (?,?,?,?,?,?)",
                    (full_name, position, phone, company_id, monthly_salary, datetime.now().strftime("%Y-%m-%d %H:%M"))
                )
                conn.commit()

        elif action == "salary":
            worker_id = int(request.form.get("worker_id", 0))
            amount = int(request.form.get("amount", 0))
            note = request.form.get("note", "").strip()

            c.execute("SELECT full_name FROM workers WHERE id=?", (worker_id,))
            w_row = c.fetchone()
            worker_name = w_row["full_name"] if w_row else "Ishchi"

            c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
            bal = c.fetchone()
            if not bal or bal["balance"] < amount:
                error = "Balans yetarli emas!"
            elif amount <= 0:
                error = "Summa 0 dan katta bo'lishi kerak!"
            else:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                c.execute(
                    "INSERT INTO salaries (worker_id, company_id, amount, note, created_at) VALUES (?,?,?,?,?)",
                    (worker_id, company_id, amount, note, now_str)
                )
                c.execute("UPDATE companies SET balance = balance - ? WHERE id=?", (amount, company_id))
                exp_desc = f"Oylik maosh — {worker_name}" + (f": {note}" if note else "")
                c.execute("INSERT INTO expenses (company_id, user_id, amount, description, created_at) VALUES (?,?,?,?,?)",
                          (company_id, session["user_id"], amount, exp_desc, now_str))
                conn.commit()
                # Telegram xabari
                c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
                comp = c.fetchone()
                c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
                new_bal = c.fetchone()
                tg_salary(comp["name"] if comp else "—", session["user"], worker_name, amount, note, now_str,
                          balance=new_bal["balance"] if new_bal else None)

        elif action == "delete":
            worker_id = int(request.form.get("worker_id", 0))
            c.execute("DELETE FROM workers WHERE id=? AND company_id=?", (worker_id, company_id))
            c.execute("DELETE FROM salaries WHERE worker_id=?", (worker_id,))
            conn.commit()

        conn.close()
        return redirect(url_for("workers"))

    # Ishchilar ro'yxati
    c.execute("""
        SELECT w.id, w.full_name, w.position, w.phone, w.created_at,
               IFNULL(SUM(s.amount), 0) AS total_salary
        FROM workers w
        LEFT JOIN salaries s ON s.worker_id = w.id
        WHERE w.company_id=?
        GROUP BY w.id
        ORDER BY w.id DESC
    """, (company_id,))
    worker_list = c.fetchall()

    # Kompaniya balansi
    c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
    bal_row = c.fetchone()
    balance = bal_row["balance"] if bal_row else 0

    conn.close()
    return render_template("workers.html", workers=worker_list, balance=balance, error=error)


# ================= PROFIL (DIRECTOR) =================
@app.route("/profile", methods=["GET", "POST"])
def profile():
    if "user" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE id=?", (session["user_id"],))
    user = c.fetchone()

    error = ""
    success = ""

    if request.method == "POST":
        action = request.form.get("action")

        if action == "change_password":
            old_pass = request.form.get("old_password", "")
            new_pass = request.form.get("new_password", "")
            confirm = request.form.get("confirm_password", "")

            if not check_password_hash(user["password"], old_pass):
                error = "Eski parol noto'g'ri!"
            elif len(new_pass) < 6:
                error = "Yangi parol kamida 6 ta belgidan iborat bo'lishi kerak!"
            elif new_pass != confirm:
                error = "Parollar mos kelmadi!"
            else:
                c.execute("UPDATE users SET password=? WHERE id=?",
                          (generate_password_hash(new_pass), session["user_id"]))
                conn.commit()
                success = "Parol muvaffaqiyatli yangilandi!"

    comp = None
    if user["company_id"]:
        c.execute("SELECT name FROM companies WHERE id=?", (user["company_id"],))
        comp = c.fetchone()

    if session.get("role") == "admin":
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses")
        total_exp = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM expenses")
        exp_count = c.fetchone()[0]
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',created_at)=strftime('%Y-%m','now')")
        month_exp = c.fetchone()[0]
    else:
        c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE user_id=?", (session["user_id"],))
        total_exp = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM expenses WHERE user_id=?", (session["user_id"],))
        exp_count = c.fetchone()[0]
        c.execute("""SELECT IFNULL(SUM(amount),0) FROM expenses
                     WHERE user_id=? AND strftime('%Y-%m',created_at)=strftime('%Y-%m','now')""",
                  (session["user_id"],))
        month_exp = c.fetchone()[0]

    conn.close()
    return render_template("profile_page.html",
        user=user,
        company_name=comp["name"] if comp else "—",
        total_exp=total_exp,
        exp_count=exp_count,
        month_exp=month_exp,
        error=error,
        success=success
    )


# ================= HISOBOTIM (DIRECTOR) =================
@app.route("/my-reports")
def my_reports():
    if "user" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else None

    # Oylik trend (oxirgi 6 oy)
    monthly = []
    for i in range(5, -1, -1):
        from datetime import date
        d = datetime.now()
        month = (d.month - i - 1) % 12 + 1
        year = d.year - ((d.month - i - 1) // 12)
        month_str = f"{year}-{month:02d}"
        c.execute("""SELECT IFNULL(SUM(amount),0) FROM expenses
                     WHERE company_id=? AND strftime('%Y-%m',created_at)=?""",
                  (company_id, month_str))
        monthly.append({"month": month_str, "total": c.fetchone()[0]})

    # Umumiy statistika
    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=?", (company_id,))
    total_all = c.fetchone()[0]

    c.execute("""SELECT IFNULL(SUM(amount),0) FROM expenses
                 WHERE company_id=? AND strftime('%Y-%m',created_at)=strftime('%Y-%m','now')""",
              (company_id,))
    total_month = c.fetchone()[0]

    c.execute("""SELECT IFNULL(SUM(amount),0) FROM expenses
                 WHERE company_id=? AND date(created_at)=date('now')""", (company_id,))
    total_today = c.fetchone()[0]

    c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
    bal = c.fetchone()
    balance = bal["balance"] if bal else 0

    # Top xarajatlar (tavsif bo'yicha)
    c.execute("""SELECT description, SUM(amount) as total, COUNT(*) as cnt
                 FROM expenses WHERE company_id=?
                 GROUP BY description ORDER BY total DESC LIMIT 8""", (company_id,))
    top_expenses = c.fetchall()

    # So'nggi 20 xarajat
    c.execute("""SELECT amount, description, created_at FROM expenses
                 WHERE company_id=? ORDER BY created_at DESC LIMIT 20""", (company_id,))
    recent = c.fetchall()

    conn.close()
    return render_template("my_reports.html",
        monthly=monthly,
        total_all=total_all,
        total_month=total_month,
        total_today=total_today,
        balance=balance,
        top_expenses=top_expenses,
        recent=recent,
    )


@app.route('/ping')
def ping():
    return 'pong', 200


# ================= GUL TAHRIRI API =================
@app.route('/api/flowers/day-data')
def flower_day_data():
    """Biror kun uchun barcha terim ma'lumotlarini qaytaradi"""
    from flask import jsonify
    if "user" not in session:
        return jsonify({'error': 'login required'}), 401

    day = request.args.get('day', '')
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else None

    c.execute("""
        SELECT h.id, h.flower_id, h.size_id, h.quantity, h.note,
               f.name as flower_name, fs.size_name
        FROM flower_harvests h
        JOIN flowers f ON h.flower_id = f.id
        LEFT JOIN flower_sizes fs ON h.size_id = fs.id
        WHERE h.company_id=? AND date(h.created_at)=?
        ORDER BY f.name, fs.size_name
    """, (company_id, day))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route('/api/flowers/update-harvest', methods=['POST'])
def update_harvest():
    """Terim sonini yangilash yoki o'chirish"""
    from flask import jsonify
    if "user" not in session:
        return jsonify({'error': 'login required'}), 401

    data = request.get_json()
    harvest_id = data.get('id')
    new_qty = int(data.get('quantity', 0))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else None

    if new_qty <= 0:
        c.execute("DELETE FROM flower_harvests WHERE id=? AND company_id=?",
                  (harvest_id, company_id))
    else:
        c.execute("UPDATE flower_harvests SET quantity=? WHERE id=? AND company_id=?",
                  (new_qty, harvest_id, company_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})



# ================= GULLAR (DIRECTOR) =================
@app.route("/flowers", methods=["GET", "POST"])
def flowers():
    if "user" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else None

    error = ""
    success = ""

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_flower":
            name = request.form.get("name", "").strip()
            sizes_hidden = request.form.get("sizes", "").strip()
            if name:
                c.execute("INSERT INTO flowers (company_id, name, created_at) VALUES (?,?,?)",
                          (company_id, name, datetime.now().strftime("%Y-%m-%d %H:%M")))
                flower_id = c.lastrowid
                for sz in [s.strip() for s in sizes_hidden.split(",") if s.strip()]:
                    c.execute("INSERT INTO flower_sizes (flower_id, size_name, created_at) VALUES (?,?,?)",
                              (flower_id, sz, datetime.now().strftime("%Y-%m-%d %H:%M")))
                conn.commit()
                success = f"\u2705 \'{name}\' guli qo\u2019shildi!"

        elif action == "bulk_harvest":
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            total_saved = 0
            for key, val in request.form.items():
                if not key.startswith("qty_"):
                    continue
                parts = key.split("_", 3)
                if len(parts) != 4:
                    continue
                _, flower_id, sort_num, size_name = parts
                quantity = int(val or 0)
                if quantity <= 0:
                    continue
                # 2-sort uchun size_id yo'q, note ga "2-sort" yozamiz
                if size_name == "2sort":
                    size_id = None
                    note = "2-sort"
                else:
                    c.execute("SELECT id FROM flower_sizes WHERE flower_id=? AND size_name=?",
                              (flower_id, size_name))
                    sz_row = c.fetchone()
                    size_id = sz_row["id"] if sz_row else None
                    note = "1-sort"
                c.execute("""INSERT INTO flower_harvests
                             (flower_id, size_id, company_id, quantity, note, created_at)
                             VALUES (?,?,?,?,?,?)""",
                          (flower_id, size_id, company_id, quantity, note, now_str))
                total_saved += quantity
            if total_saved > 0:
                conn.commit()
                success = f"✅ Jami {total_saved:,} ta gul saqlandi!"
            else:
                error = "Hech qanday son kiritilmadi!"

        elif action == "delete_flower":
            flower_id = request.form.get("flower_id")
            c.execute("DELETE FROM flower_harvests WHERE flower_id=?", (flower_id,))
            c.execute("DELETE FROM flower_sizes WHERE flower_id=?", (flower_id,))
            c.execute("DELETE FROM flowers WHERE id=? AND company_id=?", (flower_id, company_id))
            conn.commit()
            success = "Gul o\u2019chirildi"

        conn.close()
        return redirect(url_for("flowers"))

    c.execute("""
        SELECT f.id, f.name FROM flowers f
        WHERE f.company_id=? ORDER BY f.name
    """, (company_id,))
    flowers_list = c.fetchall()

    all_sizes_set = []
    size_map = {}
    for fl in flowers_list:
        c.execute("SELECT id, size_name FROM flower_sizes WHERE flower_id=? ORDER BY CAST(size_name AS INTEGER)",
                  (fl["id"],))
        for row in c.fetchall():
            sname = row["size_name"]
            if sname not in all_sizes_set:
                all_sizes_set.append(sname)
            size_map[(fl["id"], sname)] = row["id"]

    def size_sort_key(s):
        try:
            return int(''.join(filter(str.isdigit, s)))
        except:
            return 9999

    all_sizes_set.sort(key=size_sort_key)

    # Kunlik guruhlab — tarix uchun
    c.execute("""
        SELECT h.quantity, h.note as sort_info, h.created_at,
               f.name as flower_name, fs.size_name,
               date(h.created_at) as day,
               f.id as flower_id, h.id as harvest_id
        FROM flower_harvests h
        JOIN flowers f ON h.flower_id = f.id
        LEFT JOIN flower_sizes fs ON h.size_id = fs.id
        WHERE h.company_id=?
        ORDER BY h.created_at DESC LIMIT 200
    """, (company_id,))
    all_harvests = c.fetchall()

    # Kunlar bo'yicha guruhlash
    from collections import OrderedDict
    days_data = OrderedDict()
    for h in all_harvests:
        day = h["day"]
        if day not in days_data:
            days_data[day] = {"total": 0, "flowers": OrderedDict()}
        fname = h["flower_name"]
        if fname not in days_data[day]["flowers"]:
            days_data[day]["flowers"][fname] = {
                "total_s1": 0, "total_s2": 0, "sizes": []
            }
        qty = h["quantity"]
        days_data[day]["total"] += qty
        is_s2 = h["sort_info"] == "2-sort"
        if is_s2:
            days_data[day]["flowers"][fname]["total_s2"] += qty
        else:
            days_data[day]["flowers"][fname]["total_s1"] += qty
            days_data[day]["flowers"][fname]["sizes"].append({
                "size": h["size_name"] or "—",
                "qty": qty
            })

    # So'nggi 50 ta (jadval uchun ham)
    recent_harvests = all_harvests[:50]

    conn.close()
    return render_template("flowers.html",
        flowers=flowers_list,
        all_sizes=all_sizes_set,
        size_map=size_map,
        recent_harvests=recent_harvests,
        days_data=days_data,
        error=error,
        success=success
    )

def flower_sizes_api(flower_id):
    from flask import jsonify
    if "user" not in session:
        return jsonify([])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, size_name FROM flower_sizes WHERE flower_id=?", (flower_id,))
    rows = [{"id": r["id"], "name": r["size_name"]} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


# ================= GULLAR ADMIN =================
@app.route("/admin/flowers")
def admin_flowers():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    conn = get_db()
    c = conn.cursor()

    # Kompaniya + gul bo'yicha umumiy (sort ajratilgan)
    c.execute("""
        SELECT c.name as company, f.name as flower,
               IFNULL(SUM(CASE WHEN h.note!='2-sort' THEN h.quantity ELSE 0 END),0) as s1_total,
               IFNULL(SUM(CASE WHEN h.note='2-sort' THEN h.quantity ELSE 0 END),0) as s2_total,
               IFNULL(SUM(h.quantity),0) as grand_total,
               c.id as company_id
        FROM companies c
        JOIN flowers f ON f.company_id = c.id
        LEFT JOIN flower_harvests h ON h.flower_id = f.id
        GROUP BY c.id, f.id
        ORDER BY c.name, grand_total DESC
    """)
    summary = c.fetchall()

    # Razmer bo'yicha (faqat 1-sort)
    c.execute("""
        SELECT c.name as company, f.name as flower,
               IFNULL(fs.size_name, '—') as size,
               IFNULL(SUM(h.quantity), 0) as qty
        FROM companies c
        JOIN flowers f ON f.company_id = c.id
        LEFT JOIN flower_sizes fs ON fs.flower_id = f.id
        LEFT JOIN flower_harvests h ON h.flower_id = f.id
            AND h.size_id = fs.id AND h.note != '2-sort'
        GROUP BY c.id, f.id, fs.id
        HAVING qty > 0
        ORDER BY c.name, f.name, CAST(fs.size_name AS INTEGER)
    """)
    by_size = c.fetchall()

    # Oxirgi 7 kunlik trend (har kun jami)
    c.execute("""
        SELECT date(h.created_at) as day,
               IFNULL(SUM(CASE WHEN h.note!='2-sort' THEN h.quantity ELSE 0 END),0) as s1,
               IFNULL(SUM(CASE WHEN h.note='2-sort' THEN h.quantity ELSE 0 END),0) as s2,
               IFNULL(SUM(h.quantity),0) as total
        FROM flower_harvests h
        GROUP BY day ORDER BY day DESC LIMIT 14
    """)
    daily_trend = list(reversed(c.fetchall()))

    # So'nggi havestlar — kunlik guruhlab
    c.execute("""
        SELECT h.quantity, h.note as sort_info, h.created_at,
               f.name as flower, fs.size_name as size,
               c.name as company, date(h.created_at) as day
        FROM flower_harvests h
        JOIN flowers f ON h.flower_id = f.id
        JOIN companies c ON h.company_id = c.id
        LEFT JOIN flower_sizes fs ON h.size_id = fs.id
        ORDER BY h.created_at DESC LIMIT 100
    """)
    all_recent = c.fetchall()

    # Kunlar bo'yicha guruhlash
    from collections import OrderedDict
    days_summary = OrderedDict()
    for h in all_recent:
        day = h["day"]
        if day not in days_summary:
            days_summary[day] = {"total": 0, "s1": 0, "s2": 0, "companies": {}}
        days_summary[day]["total"] += h["quantity"]
        if h["sort_info"] == "2-sort":
            days_summary[day]["s2"] += h["quantity"]
        else:
            days_summary[day]["s1"] += h["quantity"]
        comp = h["company"]
        if comp not in days_summary[day]["companies"]:
            days_summary[day]["companies"][comp] = {"total": 0, "flowers": {}}
        fl = h["flower"]
        if fl not in days_summary[day]["companies"][comp]["flowers"]:
            days_summary[day]["companies"][comp]["flowers"][fl] = {"s1": [], "s2": 0}
        if h["sort_info"] == "2-sort":
            days_summary[day]["companies"][comp]["flowers"][fl]["s2"] += h["quantity"]
        else:
            days_summary[day]["companies"][comp]["flowers"][fl]["s1"].append({
                "size": h["size"] or "—", "qty": h["quantity"]
            })
        days_summary[day]["companies"][comp]["total"] += h["quantity"]

    # Umumiy statistika
    c.execute("SELECT IFNULL(SUM(quantity),0) FROM flower_harvests")
    total_all = c.fetchone()[0]
    c.execute("SELECT IFNULL(SUM(quantity),0) FROM flower_harvests WHERE date(created_at)=date('now')")
    total_today = c.fetchone()[0]
    c.execute("""SELECT IFNULL(SUM(quantity),0) FROM flower_harvests
                 WHERE strftime('%Y-%m',created_at)=strftime('%Y-%m','now')""")
    total_month = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT date(created_at)) FROM flower_harvests")
    total_days = c.fetchone()[0]

    c.execute("SELECT id, name FROM companies ORDER BY name")
    companies = c.fetchall()

    conn.close()
    return render_template("admin_flowers.html",
        summary=summary,
        by_size=by_size,
        daily_trend=daily_trend,
        days_summary=days_summary,
        total_all=total_all,
        total_today=total_today,
        total_month=total_month,
        total_days=total_days,
        companies=companies
    )



@app.route('/sw.js')
def service_worker():
    from flask import send_from_directory
    return send_from_directory('static', 'sw.js',
                               mimetype='application/javascript')


@app.route('/offline')
def offline():
    return '''<!DOCTYPE html><html><head><meta charset="UTF-8">
    <meta name="viewport" content="width=device-width,initial-scale=1">
    <title>Offline | Company Accounting</title>
    <style>body{font-family:sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;background:#f0f2f8;margin:0}
    .box{text-align:center;padding:40px;background:white;border-radius:20px;box-shadow:0 4px 24px rgba(0,0,0,.08)}
    .emoji{font-size:60px;margin-bottom:16px}.title{font-size:22px;font-weight:700;color:#0f172a;margin-bottom:8px}
    .sub{color:#64748b;font-size:14px;margin-bottom:24px}
    .btn{background:linear-gradient(135deg,#3b6ef6,#00d4aa);color:white;border:none;border-radius:12px;padding:12px 28px;font-size:15px;font-weight:600;cursor:pointer}
    </style></head><body>
    <div class="box">
      <div class="emoji">📡</div>
      <div class="title">Internet yo'q</div>
      <div class="sub">Internetga ulanib qayta urinib ko'ring</div>
      <button class="btn" onclick="location.reload()">🔄 Qayta urinish</button>
    </div></body></html>''', 200


@app.route('/admin/send-daily-report')
def send_daily_report_now():
    if session.get('role') != 'admin':
        return "Ruxsat yo'q ❌"
    try:
        import daily_report
        daily_report.main()
        return "✅ Hisobot muvaffaqiyatli yuborildi! Telegramni tekshiring."
    except Exception as e:
        return f"❌ Xatolik: {e}"


@app.route('/admin/send-report-telegram', methods=['POST'])
def send_report_telegram():
    if session.get('role') != 'admin':
        return {'error': 'ruxsat yoq'}, 401
    from flask import jsonify

    date_from = request.form.get('date_from', '')
    date_to = request.form.get('date_to', '')
    company_id = request.form.get('company_id', 'all')
    fmt = request.form.get('format', 'both')

    try:
        conn = get_db()
        c = conn.cursor()

        # Sana sharti
        if date_from and date_to:
            date_cond = f"date(e.created_at) BETWEEN '{date_from}' AND '{date_to}'"
            period_label = f"{date_from} — {date_to}"
        elif date_from:
            date_cond = f"date(e.created_at) >= '{date_from}'"
            period_label = f"{date_from} dan"
        else:
            date_cond = "1=1"
            period_label = "Barcha vaqt"

        # Kompaniya sharti
        if company_id != 'all':
            comp_cond = f"e.company_id = {int(company_id)}"
            c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
            row = c.fetchone()
            comp_label = row['name'] if row else "Noma'lum"
        else:
            comp_cond = "1=1"
            comp_label = "Barcha kompaniyalar"

        where = f"WHERE {date_cond} AND {comp_cond}"

        # Xarajatlar
        c.execute(f"""
            SELECT e.amount, e.description, e.created_at,
                   c.name as company, u.username as director
            FROM expenses e
            JOIN companies c ON e.company_id=c.id
            LEFT JOIN users u ON e.user_id=u.id
            {where}
            ORDER BY e.created_at DESC
        """)
        expenses = c.fetchall()

        # Kompaniyalar balansi
        if company_id != 'all':
            c.execute("SELECT id, name, balance FROM companies WHERE id=?", (company_id,))
        else:
            c.execute("SELECT id, name, balance FROM companies ORDER BY balance DESC")
        companies = c.fetchall()

        total_exp = sum(e['amount'] for e in expenses)
        conn.close()

        # Telegram xabar
        msg = (
            f"╔══════════════════════╗\n"
            f"║  📊  <b>MAXSUS HISOBOT</b>  📊  ║\n"
            f"╚══════════════════════╝\n\n"
            f"📅 <b>Davr:</b> {period_label}\n"
            f"🏢 <b>Kompaniya:</b> {comp_label}\n"
            f"💸 <b>Jami xarajat:</b> <code>{total_exp:,} so'm</code>\n"
            f"📋 <b>Xarajatlar soni:</b> {len(expenses)} ta\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🏦 <b>BALANSLAR</b>\n"
        )
        for comp in companies:
            emoji = "🟢" if comp['balance'] > 0 else "🔴"
            msg += f"{emoji} {comp['name']}: <code>{comp['balance']:,} so'm</code>\n"

        send_telegram_message(msg)

        # Excel
        if fmt in ('excel', 'both'):
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Hisobot"
            ws.sheet_view.showGridLines = False

            # Sarlavha
            ws.merge_cells("A1:F1")
            ws["A1"] = f"HISOBOT: {period_label} | {comp_label}"
            ws["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
            ws["A1"].fill = PatternFill("solid", fgColor="0F2044")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Jadval sarlavhasi
            heads = ["#", "Sana", "Kompaniya", "Direktor", "Summa (so'm)", "Izoh"]
            widths = [5, 18, 22, 16, 18, 35]
            side = Side(style="thin", color="D0D7E3")
            b = Border(left=side, right=side, top=side, bottom=side)
            for j, (h, w) in enumerate(zip(heads, widths), 1):
                cell = ws.cell(row=2, column=j, value=h)
                cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
                cell.fill = PatternFill("solid", fgColor="1E3A5F")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = b
                ws.column_dimensions[get_column_letter(j)].width = w
            ws.row_dimensions[2].height = 20

            for i, exp in enumerate(expenses, 1):
                row = i + 2
                bg = "F0F7FF" if i % 2 == 0 else "FFFFFF"
                vals = [i, str(exp['created_at'])[:16], exp['company'],
                        exp['director'] or '—', exp['amount'], exp['description'] or '—']
                for j, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=j, value=v)
                    cell.font = Font(size=9, name="Arial")
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(horizontal="right" if j == 5 else "center" if j in [1,2,4] else "left")
                    cell.border = b
                    if j == 5:
                        cell.number_format = '#,##0'
                ws.row_dimensions[row].height = 16

            # Jami
            tr = len(expenses) + 3
            ws.cell(row=tr, column=4, value="JAMI:").font = Font(bold=True, size=10, name="Arial")
            tc = ws.cell(row=tr, column=5, value=total_exp)
            tc.font = Font(bold=True, color="DC2626", size=11, name="Arial")
            tc.number_format = '#,##0'
            tc.alignment = Alignment(horizontal="right")
            for ci in range(1, 7):
                ws.cell(row=tr, column=ci).fill = PatternFill("solid", fgColor="E8F4FD")
                ws.cell(row=tr, column=ci).border = b
            ws.row_dimensions[tr].height = 22

            import io as _io
            xbuf = _io.BytesIO()
            wb.save(xbuf)
            xbuf.seek(0)
            fname = f"hisobot_{date_from or 'barchasi'}_{comp_label[:15]}.xlsx"
            send_telegram_file_direct(xbuf.read(), fname, f"📊 Excel | {period_label} | {comp_label}")

        # PDF
        if fmt in ('pdf', 'both'):
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                             Paragraph, Spacer, HRFlowable)
            import io as _io

            pbuf = _io.BytesIO()
            doc = SimpleDocTemplate(pbuf, pagesize=A4,
                                    rightMargin=1.5*cm, leftMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            story = []

            story.append(Paragraph(f"Hisobot: {period_label}",
                ParagraphStyle('t', parent=styles['Normal'], fontSize=16,
                               fontName='Helvetica-Bold',
                               textColor=rl_colors.HexColor('#0f2044'), spaceAfter=3)))
            story.append(Paragraph(f"{comp_label}  |  Yaratildi: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                ParagraphStyle('s', parent=styles['Normal'], fontSize=9,
                               textColor=rl_colors.HexColor('#64748b'), spaceAfter=10)))
            story.append(HRFlowable(width="100%", thickness=2,
                                    color=rl_colors.HexColor('#3b6ef6')))
            story.append(Spacer(1, 10))

            # Balanslar
            story.append(Paragraph("Kompaniyalar Balansi",
                ParagraphStyle('h2', parent=styles['Normal'], fontSize=11,
                               fontName='Helvetica-Bold', spaceAfter=6)))
            bal_data = [["Kompaniya", "Balans"]]
            for comp in companies:
                bal_data.append([comp['name'], f"{comp['balance']:,} so'm"])
            bt = Table(bal_data, colWidths=[10*cm, 8*cm])
            bt.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), rl_colors.HexColor('#0f2044')),
                ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [rl_colors.HexColor('#f0f7ff'), rl_colors.white]),
                ('GRID', (0,0), (-1,-1), 0.4, rl_colors.HexColor('#dee4f0')),
                ('ALIGN', (1,0), (1,-1), 'RIGHT'),
                ('PADDING', (0,0), (-1,-1), 7),
                ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ]))
            story.append(bt)
            story.append(Spacer(1, 12))

            # Xarajatlar
            story.append(Paragraph(f"Xarajatlar ({len(expenses)} ta, jami: {total_exp:,} so'm)",
                ParagraphStyle('h2', parent=styles['Normal'], fontSize=11,
                               fontName='Helvetica-Bold', spaceAfter=6)))
            exp_data = [["#", "Sana", "Kompaniya", "Direktor", "Summa", "Izoh"]]
            for i, exp in enumerate(expenses, 1):
                desc = str(exp['description'] or '')[:30]
                exp_data.append([str(i), str(exp['created_at'])[:16],
                                  exp['company'], exp['director'] or '—',
                                  f"{exp['amount']:,}", desc])
            exp_data.append(["", "", "", "JAMI", f"{total_exp:,}", ""])
            et = Table(exp_data, colWidths=[0.8*cm, 3.2*cm, 3.5*cm, 3*cm, 3*cm, 4.5*cm])
            et.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), rl_colors.HexColor('#7f1d1d')),
                ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ROWBACKGROUNDS', (0,1), (-1,-2), [rl_colors.HexColor('#fff5f5'), rl_colors.white]),
                ('BACKGROUND', (0,-1), (-1,-1), rl_colors.HexColor('#ffe4e4')),
                ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (4,-1), (4,-1), rl_colors.HexColor('#dc2626')),
                ('GRID', (0,0), (-1,-1), 0.4, rl_colors.HexColor('#fecaca')),
                ('ALIGN', (4,0), (4,-1), 'RIGHT'),
                ('ALIGN', (0,0), (0,-1), 'CENTER'),
                ('PADDING', (0,0), (-1,-1), 5),
                ('FONTNAME', (0,1), (-1,-2), 'Helvetica'),
            ]))
            story.append(et)
            doc.build(story)
            pbuf.seek(0)
            fname = f"hisobot_{date_from or 'barchasi'}_{comp_label[:15]}.pdf"
            send_telegram_file_direct(pbuf.read(), fname, f"📄 PDF | {period_label} | {comp_label}")

        return jsonify({'ok': True, 'message': f"✅ Telegram ga yuborildi! ({len(expenses)} ta xarajat)"})

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'message': f"❌ Xatolik: {str(e)}"})


def send_telegram_file_direct(file_bytes, filename, caption):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
            files={"document": (filename, file_bytes)},
            timeout=30
        )
    except Exception as e:
        print("Fayl yuborish xato:", e)


# ================= PDF HISOBOTLAR =================
@app.route("/admin/export/pdf/all")
def export_pdf_all():
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    conn = get_db()
    c = conn.cursor()

    c.execute("""
        SELECT c.name, c.balance, IFNULL(SUM(e.amount),0) as total_exp,
               u.username as director
        FROM companies c
        LEFT JOIN expenses e ON e.company_id=c.id
        LEFT JOIN users u ON u.company_id=c.id AND u.role='director'
        GROUP BY c.id ORDER BY c.id DESC
    """)
    companies = c.fetchall()

    c.execute("""
        SELECT e.description, e.amount, e.created_at, c.name, u.username, e.is_edited
        FROM expenses e
        JOIN companies c ON e.company_id=c.id
        LEFT JOIN users u ON e.user_id=u.id
        ORDER BY e.created_at DESC LIMIT 100
    """)
    expenses = c.fetchall()

    c.execute("SELECT IFNULL(SUM(balance),0) FROM companies")
    total_balance = c.fetchone()[0]
    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses")
    total_expenses = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM companies")
    total_companies = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM users WHERE role='director'")
    total_directors = c.fetchone()[0]

    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    story = []

    # SARLAVHA
    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                  fontSize=20, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1e293b'),
                                  spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=10, fontName='Helvetica',
                                textColor=colors.HexColor('#64748b'),
                                spaceAfter=16)

    story.append(Paragraph("Company Accounting - Umumiy Hisobot", title_style))
    story.append(Paragraph(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#3b6ef6')))
    story.append(Spacer(1, 14))

    # UMUMIY STATISTIKA
    story.append(Paragraph("Umumiy Ko'rsatkichlar", ParagraphStyle('h2', parent=styles['Normal'],
                  fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'), spaceAfter=8)))

    stat_data = [
        ["Ko'rsatkich", "Qiymat"],
        ["Umumiy balans", f"{total_balance:,} so'm"],
        ["Jami xarajatlar", f"{total_expenses:,} so'm"],
        ["Kompaniyalar soni", str(total_companies)],
        ["Direktorlar soni", str(total_directors)],
    ]
    stat_table = Table(stat_data, colWidths=[8*cm, 8*cm])
    stat_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b6ef6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ]))
    story.append(stat_table)
    story.append(Spacer(1, 18))

    # KOMPANIYALAR JADVALI
    story.append(Paragraph("Kompaniyalar", ParagraphStyle('h2', parent=styles['Normal'],
                  fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'), spaceAfter=8)))

    comp_data = [["#", "Kompaniya", "Direktor", "Balans", "Xarajat"]]
    for i, row in enumerate(companies, 1):
        comp_data.append([
            str(i), str(row[0] or "-"), str(row[3] or "-"),
            f"{row[1]:,}", f"{row[2]:,}"
        ])

    comp_table = Table(comp_data, colWidths=[1*cm, 5*cm, 4*cm, 4*cm, 4*cm])
    comp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (3,0), (4,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 7),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ]))
    story.append(comp_table)
    story.append(Spacer(1, 18))

    # SO'NGI XARAJATLAR
    story.append(Paragraph("So'nggi 100 ta xarajat", ParagraphStyle('h2', parent=styles['Normal'],
                  fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'), spaceAfter=8)))

    exp_data = [["#", "Sana", "Kompaniya", "Direktor", "Summa", "Izoh"]]
    for i, row in enumerate(expenses, 1):
        desc = str(row[0] or "")[:30] + ("..." if len(str(row[0] or "")) > 30 else "")
        edited = " (*)" if row[5] else ""
        exp_data.append([
            str(i), str(row[2] or "")[:16],
            str(row[3] or "")[:15], str(row[4] or "-")[:12],
            f"{row[1]:,}", desc + edited
        ])

    exp_table = Table(exp_data, colWidths=[0.8*cm, 2.8*cm, 3.5*cm, 2.8*cm, 2.8*cm, 5.3*cm])
    exp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#fff5f5'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (4,0), (4,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 5),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ]))
    story.append(exp_table)

    story.append(Spacer(1, 10))
    story.append(Paragraph("(*) — Tahrirlangan xarajat", ParagraphStyle('note', parent=styles['Normal'],
                  fontSize=8, textColor=colors.HexColor('#94a3b8'))))

    doc.build(story)
    buf.seek(0)

    fname = f"hisobot_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return Response(buf.read(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/admin/export/pdf/company/<int:company_id>")
def export_pdf_company(company_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    import io

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM companies WHERE id=?", (company_id,))
    company = c.fetchone()
    if not company:
        conn.close()
        return "Kompaniya topilmadi", 404

    c.execute("""
        SELECT u.username FROM users u
        WHERE u.company_id=? AND u.role='director'
    """, (company_id,))
    directors = [r[0] for r in c.fetchall()]

    c.execute("""
        SELECT e.description, e.amount, e.created_at, u.username, e.is_edited
        FROM expenses e
        LEFT JOIN users u ON e.user_id=u.id
        WHERE e.company_id=?
        ORDER BY e.created_at DESC
    """, (company_id,))
    expenses = c.fetchall()

    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE company_id=?", (company_id,))
    total_exp = c.fetchone()[0]

    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"{company['name']} - Hisobot", ParagraphStyle('title', parent=styles['Normal'],
                  fontSize=18, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e293b'), spaceAfter=4)))
    story.append(Paragraph(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ParagraphStyle('sub',
                  parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=14)))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#3b6ef6')))
    story.append(Spacer(1, 14))

    stat_data = [
        ["Ko'rsatkich", "Qiymat"],
        ["Joriy balans", f"{company['balance']:,} so'm"],
        ["Jami xarajatlar", f"{total_exp:,} so'm"],
        ["Direktorlar", ", ".join(directors) or "-"],
        ["Yaratilgan", str(company['created_at'] or "-")],
    ]
    stat_table = Table(stat_data, colWidths=[7*cm, 10*cm])
    stat_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b6ef6')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ]))
    story.append(stat_table)
    story.append(Spacer(1, 18))

    story.append(Paragraph("Barcha xarajatlar", ParagraphStyle('h2', parent=styles['Normal'],
                  fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a'), spaceAfter=8)))

    exp_data = [["#", "Sana", "Direktor", "Summa", "Izoh"]]
    for i, row in enumerate(expenses, 1):
        desc = str(row[0] or "")[:35] + ("..." if len(str(row[0] or "")) > 35 else "")
        exp_data.append([str(i), str(row[2] or "")[:16], str(row[3] or "-"), f"{row[1]:,}", desc])

    exp_table = Table(exp_data, colWidths=[0.8*cm, 3.5*cm, 3.5*cm, 3.5*cm, 6.7*cm])
    exp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (3,0), (3,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 5),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
    ]))
    story.append(exp_table)

    doc.build(story)
    buf.seek(0)

    fname = f"{company['name'].replace(' ', '_')}_hisobot_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.read(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


# ================= WORKER DETAIL =================
@app.route("/workers/<int:worker_id>", methods=["GET", "POST"])
def worker_detail(worker_id):
    if "user" not in session:
        return redirect(url_for("login"))

    conn = get_db()
    c = conn.cursor()

    # company_id ni DB dan olamiz — session da bo'lmasa ham ishlaydi
    c.execute("SELECT company_id FROM users WHERE id=?", (session["user_id"],))
    u = c.fetchone()
    company_id = u["company_id"] if u else session.get("company_id")

    c.execute("SELECT * FROM workers WHERE id=? AND company_id=?", (worker_id, company_id))
    worker = c.fetchone()
    if not worker:
        conn.close()
        return "Ishchi topilmadi", 404

    error = ""
    success = ""

    if request.method == "POST":
        action = request.form.get("action")

        if action == "toggle_status":
            new_status = "paused" if worker["status"] == "active" else "active"
            c.execute("UPDATE workers SET status=? WHERE id=?", (new_status, worker_id))
            conn.commit()

        elif action == "update_salary":
            new_salary = int(request.form.get("monthly_salary", 0) or 0)
            c.execute("UPDATE workers SET monthly_salary=? WHERE id=?", (new_salary, worker_id))
            conn.commit()
            success = "Oylik maosh yangilandi!"

        elif action == "attendance":
            date = request.form.get("date")
            att_status = request.form.get("att_status", "absent")
            note = request.form.get("note", "")
            c.execute("SELECT id FROM attendance WHERE worker_id=? AND date=?", (worker_id, date))
            if c.fetchone():
                c.execute("UPDATE attendance SET status=?, note=? WHERE worker_id=? AND date=?",
                          (att_status, note, worker_id, date))
            else:
                c.execute("INSERT INTO attendance (worker_id, company_id, date, status, note, created_at) VALUES (?,?,?,?,?,?)",
                          (worker_id, company_id, date, att_status, note, datetime.now().strftime("%Y-%m-%d %H:%M")))
            conn.commit()

        elif action == "advance":
            amount = int(request.form.get("amount", 0) or 0)
            note = request.form.get("note", "").strip()
            c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
            bal = c.fetchone()
            if not bal or bal["balance"] < amount:
                error = "Balans yetarli emas!"
            elif amount <= 0:
                error = "Summa 0 dan katta bo'lishi kerak!"
            else:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                c.execute("INSERT INTO advances (worker_id, company_id, amount, note, created_at) VALUES (?,?,?,?,?)",
                          (worker_id, company_id, amount, note, now_str))
                c.execute("UPDATE companies SET balance = balance - ? WHERE id=?", (amount, company_id))
                exp_desc = f"Avans — {worker['full_name']}" + (f": {note}" if note else "")
                c.execute("INSERT INTO expenses (company_id, user_id, amount, description, created_at) VALUES (?,?,?,?,?)",
                          (company_id, session["user_id"], amount, exp_desc, now_str))
                conn.commit()
                # Telegram xabari
                c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
                comp = c.fetchone()
                c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
                new_bal = c.fetchone()
                tg_advance(comp["name"] if comp else "—", session["user"], worker["full_name"], amount, note, now_str,
                           balance=new_bal["balance"] if new_bal else None)
                success = "Avans berildi!"

        elif action == "pay_salary":
            amount = int(request.form.get("amount", 0) or 0)
            note = request.form.get("note", "").strip()
            c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
            bal = c.fetchone()
            if not bal or bal["balance"] < amount:
                error = "Balans yetarli emas!"
            elif amount <= 0:
                error = "Summa 0 dan katta bo'lishi kerak!"
            else:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                c.execute("INSERT INTO salaries (worker_id, company_id, amount, note, created_at) VALUES (?,?,?,?,?)",
                          (worker_id, company_id, amount, note, now_str))
                c.execute("UPDATE companies SET balance = balance - ? WHERE id=?", (amount, company_id))
                exp_desc = f"Oylik maosh — {worker['full_name']}" + (f": {note}" if note else "")
                c.execute("INSERT INTO expenses (company_id, user_id, amount, description, created_at) VALUES (?,?,?,?,?)",
                          (company_id, session["user_id"], amount, exp_desc, now_str))
                conn.commit()
                # Telegram xabari
                c.execute("SELECT name FROM companies WHERE id=?", (company_id,))
                comp = c.fetchone()
                c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
                new_bal = c.fetchone()
                tg_salary(comp["name"] if comp else "—", session["user"], worker["full_name"], amount, note, now_str,
                          balance=new_bal["balance"] if new_bal else None)
                success = "Maosh to'landi!"

        conn.close()
        return redirect(url_for("worker_detail", worker_id=worker_id))

    now = datetime.now()
    month_str = now.strftime("%Y-%m")

    c.execute("SELECT IFNULL(SUM(amount),0) FROM salaries WHERE worker_id=? AND created_at LIKE ?",
              (worker_id, month_str + "%"))
    month_salary_paid = c.fetchone()[0]

    c.execute("SELECT IFNULL(SUM(amount),0) FROM advances WHERE worker_id=? AND created_at LIKE ?",
              (worker_id, month_str + "%"))
    month_advances = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM attendance WHERE worker_id=? AND status='present' AND date LIKE ?",
              (worker_id, month_str + "%"))
    days_present = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM attendance WHERE worker_id=? AND status='absent' AND date LIKE ?",
              (worker_id, month_str + "%"))
    days_absent = c.fetchone()[0]

    monthly_salary = worker["monthly_salary"] or 0
    salary_left = monthly_salary - month_salary_paid - month_advances

    c.execute("SELECT date, status, note FROM attendance WHERE worker_id=? ORDER BY date DESC LIMIT 30",
              (worker_id,))
    attendance_list = c.fetchall()

    c.execute("SELECT amount, note, created_at FROM salaries WHERE worker_id=? ORDER BY created_at DESC LIMIT 10",
              (worker_id,))
    salary_history = c.fetchall()

    c.execute("SELECT amount, note, created_at FROM advances WHERE worker_id=? ORDER BY created_at DESC LIMIT 10",
              (worker_id,))
    advance_history = c.fetchall()

    c.execute("SELECT balance FROM companies WHERE id=?", (company_id,))
    bal_row = c.fetchone()
    balance = bal_row["balance"] if bal_row else 0

    conn.close()
    return render_template("worker_detail.html",
        worker=worker,
        month_salary_paid=month_salary_paid,
        month_advances=month_advances,
        days_present=days_present,
        days_absent=days_absent,
        salary_left=salary_left,
        attendance_list=attendance_list,
        salary_history=salary_history,
        advance_history=advance_history,
        balance=balance,
        error=error,
        success=success,
        today=now.strftime("%Y-%m-%d"),
        month_name=now.strftime("%B %Y")
    )


@app.route("/admin/delete-company/<int:company_id>", methods=["POST"])
def delete_company(company_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"
    conn = get_db()
    c = conn.cursor()
    # Kompaniyaga bog'liq direktorlarni ajratib olamiz
    c.execute("UPDATE users SET company_id=NULL WHERE company_id=?", (company_id,))
    # Kompaniya xarajatlarini o'chiramiz
    c.execute("DELETE FROM expenses WHERE company_id=?", (company_id,))
    # Kompaniyani o'chiramiz
    c.execute("DELETE FROM companies WHERE id=?", (company_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("admin_companies"))


@app.route("/admin/delete-director/<int:director_id>", methods=["POST"])
def delete_director(director_id):
    if session.get("role") != "admin":
        return "Ruxsat yo'q ❌"
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=? AND role='director'", (director_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("admin_directors"))

# ================= QOLDI MODAL API =================
@app.route('/api/company-balances')
def company_balances():
    if 'user' not in session:
        return {'error': 'login required'}, 401
    from flask import jsonify
    conn = get_db()
    c = conn.cursor()
    if session.get('role') == 'admin':
        c.execute("""
            SELECT c.id, c.name, c.balance, u.username as director
            FROM companies c
            LEFT JOIN users u ON u.company_id=c.id AND u.role='director'
            ORDER BY c.balance DESC
        """)
    else:
        c.execute("""
            SELECT c.id, c.name, c.balance, u.username as director
            FROM companies c
            LEFT JOIN users u ON u.company_id=c.id AND u.role='director'
            WHERE c.id=?
        """, (session.get('company_id'),))
    rows = [{'id': r['id'], 'name': r['name'], 'balance': r['balance'], 'director': r['director'] or '—'} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route('/api/directors')
def api_directors():
    if 'user' not in session or session.get('role') != 'admin':
        return {'error': 'ruxsat yoq'}, 401
    from flask import jsonify
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT u.id, u.username, u.created_at,
               c.name as company, c.id as company_id
        FROM users u
        LEFT JOIN companies c ON u.company_id=c.id
        WHERE u.role='director'
        ORDER BY u.id DESC
    """)
    rows = [{'id': r['id'], 'username': r['username'], 'created_at': r['created_at'] or '',
             'company': r['company'] or '—', 'company_id': r['company_id']} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route('/api/workers')
def api_workers():
    if 'user' not in session or session.get('role') != 'admin':
        return {'error': 'ruxsat yoq'}, 401
    from flask import jsonify
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT w.id, w.full_name, w.position, w.status,
               w.monthly_salary, c.name as company, c.id as company_id
        FROM workers w
        LEFT JOIN companies c ON w.company_id=c.id
        ORDER BY w.id DESC
    """)
    rows = [{'id': r['id'], 'full_name': r['full_name'], 'position': r['position'] or '—',
             'status': r['status'] or 'active', 'monthly_salary': r['monthly_salary'] or 0,
             'company': r['company'] or '—', 'company_id': r['company_id']} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


@app.route('/api/expenses-by-period')
def api_expenses_by_period():
    if 'user' not in session or session.get('role') != 'admin':
        return {'error': 'ruxsat yoq'}, 401
    from flask import jsonify
    period = request.args.get('period', 'month')
    conn = get_db()
    c = conn.cursor()
    if period == 'today':
        condition = "date(e.created_at) = date('now')"
    else:
        condition = "strftime('%Y-%m', e.created_at) = strftime('%Y-%m', 'now')"
    c.execute(f"""
        SELECT e.amount, e.description, e.created_at, c.name as company, u.username as director
        FROM expenses e
        JOIN companies c ON e.company_id = c.id
        LEFT JOIN users u ON e.user_id = u.id
        WHERE {condition}
        ORDER BY e.created_at DESC
        LIMIT 50
    """)
    rows = [{'amount': r['amount'], 'description': r['description'] or '—',
             'created_at': r['created_at'] or '', 'company': r['company'],
             'director': r['director'] or '—'} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)


# ================= RUN =================
init_db()  # Gunicorn va oddiy run ikkalasida ham ishlaydi

if __name__ == "__main__":
    app.run(debug=False)
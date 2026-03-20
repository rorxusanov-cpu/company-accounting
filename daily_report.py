import sqlite3, os, io, requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable, PageBreak)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "8588085417:AAG1_uFr9irp7-E2fGd20jg0BbxxUopSsH4")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "5703562662")
DB_PATH = "/data/users.db" if os.path.exists("/data") else "users.db"
TODAY = datetime.now().strftime("%Y-%m-%d")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
MONTH = datetime.now().strftime("%Y-%m")
MONTH_NAME = datetime.now().strftime("%B %Y")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_report_data():
    conn = get_db()
    c = conn.cursor()

    # Kompaniyalar
    c.execute("""
        SELECT c.id, c.name, c.balance,
               IFNULL((SELECT SUM(amount) FROM expenses WHERE company_id=c.id AND date(created_at)=date('now')), 0) as today_exp,
               IFNULL((SELECT SUM(amount) FROM expenses WHERE company_id=c.id AND strftime('%Y-%m',created_at)=?), 0) as month_exp,
               IFNULL((SELECT SUM(amount) FROM expenses WHERE company_id=c.id), 0) as total_exp,
               (SELECT username FROM users WHERE company_id=c.id AND role='director' LIMIT 1) as director
        FROM companies c ORDER BY c.balance DESC
    """, (MONTH,))
    companies = c.fetchall()

    # Bugungi xarajatlar
    c.execute("""
        SELECT e.amount, e.description, e.created_at, c.name as company,
               u.username as director, c.id as company_id
        FROM expenses e
        JOIN companies c ON e.company_id=c.id
        LEFT JOIN users u ON e.user_id=u.id
        WHERE date(e.created_at)=date('now')
        ORDER BY c.name, e.created_at DESC
    """)
    today_expenses = c.fetchall()

    # Umumiy raqamlar
    c.execute("SELECT IFNULL(SUM(balance),0) FROM companies")
    total_balance = c.fetchone()[0]
    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE date(created_at)=date('now')")
    total_today = c.fetchone()[0]
    c.execute("SELECT IFNULL(SUM(amount),0) FROM expenses WHERE strftime('%Y-%m',created_at)=?", (MONTH,))
    total_month = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM companies")
    companies_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM users WHERE role='director'")
    directors_count = c.fetchone()[0]
    try:
        c.execute("SELECT COUNT(*) FROM workers WHERE status='active'")
        workers_count = c.fetchone()[0]
    except:
        workers_count = 0

    # ===== GULLAR TERIMI =====
    # Bugungi jami terim
    try:
        c.execute("""
            SELECT IFNULL(SUM(quantity), 0) FROM flower_harvests
            WHERE date(created_at) = date('now')
        """)
        flowers_today_total = c.fetchone()[0]

        # Oylik terim
        c.execute("""
            SELECT IFNULL(SUM(quantity), 0) FROM flower_harvests
            WHERE strftime('%Y-%m', created_at) = ?
        """, (MONTH,))
        flowers_month_total = c.fetchone()[0]

        # Bugungi terim — gul turlari bo'yicha
        c.execute("""
            SELECT f.name as flower_name,
                   SUM(CASE WHEN h.note != '2-sort' THEN h.quantity ELSE 0 END) as s1,
                   SUM(CASE WHEN h.note = '2-sort' THEN h.quantity ELSE 0 END) as s2,
                   SUM(h.quantity) as total
            FROM flower_harvests h
            JOIN flowers f ON h.flower_id = f.id
            WHERE date(h.created_at) = date('now')
            GROUP BY f.name
            ORDER BY total DESC
        """)
        flowers_today_by_type = c.fetchall()

        # Kompaniya bo'yicha bugungi terim
        c.execute("""
            SELECT comp.name as company, SUM(h.quantity) as total
            FROM flower_harvests h
            JOIN companies comp ON h.company_id = comp.id
            WHERE date(h.created_at) = date('now')
            GROUP BY comp.name
            ORDER BY total DESC
        """)
        flowers_today_by_company = c.fetchall()

    except Exception as e:
        print("Gullar ma'lumotlari xatosi:", e)
        flowers_today_total = 0
        flowers_month_total = 0
        flowers_today_by_type = []
        flowers_today_by_company = []

    conn.close()
    return dict(
        companies=companies,
        today_expenses=today_expenses,
        total_balance=total_balance,
        total_today=total_today,
        total_month=total_month,
        companies_count=companies_count,
        directors_count=directors_count,
        workers_count=workers_count,
        flowers_today_total=flowers_today_total,
        flowers_month_total=flowers_month_total,
        flowers_today_by_type=flowers_today_by_type,
        flowers_today_by_company=flowers_today_by_company,
    )


def send_telegram_message(text):
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                      data={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"}, timeout=10)
    except Exception as e:
        print("Telegram xato:", e)


def send_telegram_file(file_bytes, filename, caption):
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
                      data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
                      files={"document": (filename, file_bytes)}, timeout=30)
    except Exception as e:
        print("Fayl yuborish xato:", e)


def send_daily_telegram(data):
    # Kompaniyalar qatori
    comp_lines = ""
    for comp in data["companies"]:
        bal_icon = "🟢" if comp["balance"] > 0 else "🔴"
        comp_lines += (
            f"\n{bal_icon} <b>{comp['name']}</b>"
            f"{'  · 👔 ' + comp['director'] if comp['director'] else ''}\n"
            f"   💰 Balans: <code>{comp['balance']:,} so'm</code>\n"
            f"   ☀️ Bugun: <code>{comp['today_exp']:,} so'm</code>\n"
            f"   📅 Oy: <code>{comp['month_exp']:,} so'm</code>\n"
        )

    # Gullar qatori
    flower_lines = ""
    if data["flowers_today_total"] > 0:
        for f in data["flowers_today_by_type"]:
            s1 = f["s1"] or 0
            s2 = f["s2"] or 0
            flower_lines += (
                f"\n🌸 <b>{f['flower_name']}</b>: <code>{f['total']:,} ta</code>"
                f"  (1-sort: {s1:,} | 2-sort: {s2:,})"
            )
        if data["flowers_today_by_company"]:
            flower_lines += "\n\n🏢 <b>Kompaniya bo'yicha:</b>"
            for fc in data["flowers_today_by_company"]:
                flower_lines += f"\n   • {fc['company']}: <code>{fc['total']:,} ta</code>"
    else:
        flower_lines = "\n   Bugun terim kiritilmagan"

    send_telegram_message(
        f"╔══════════════════════╗\n"
        f"║  📊  <b>KUNLIK HISOBOT</b>  📊  ║\n"
        f"╚══════════════════════╝\n\n"
        f"📅 <b>Sana:</b> {TODAY}    🕗 <b>Vaqt:</b> 20:00\n\n"

        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 <b>MOLIYA</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💳 Jami balans: <code>{data['total_balance']:,} so'm</code>\n"
        f"☀️ Bugungi xarajat: <code>{data['total_today']:,} so'm</code>\n"
        f"📅 Oylik xarajat: <code>{data['total_month']:,} so'm</code>\n"
        f"🏢 Kompaniyalar: <b>{data['companies_count']}</b>  "
        f"👔 Direktorlar: <b>{data['directors_count']}</b>  "
        f"👷 Ishchilar: <b>{data['workers_count']}</b>\n\n"

        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🌸 <b>GULLAR TERIMI</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"☀️ Bugun: <code>{data['flowers_today_total']:,} ta</code>    "
        f"📅 Bu oy: <code>{data['flowers_month_total']:,} ta</code>"
        f"{flower_lines}\n\n"

        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🏢 <b>KOMPANIYALAR</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
        f"{comp_lines}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💸 Bugun <b>{len(data['today_expenses'])}</b> ta xarajat\n"
        f"📎 Excel va PDF quyida 👇"
    )


# ===== STIL YORDAMCHILARI =====
def hdr(text, ws, row, col, end_col, bg="1E3A5F", fg="FFFFFF", size=11, height=22):
    ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(end_col)}{row}")
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=size, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height
    return cell

def styled(cell, bold=False, color="000000", bg=None, align="left", size=10, wrap=False, fmt=None):
    cell.font = Font(bold=bold, color=color, size=size, name="Arial")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    return cell

def border_all(ws, min_row, max_row, min_col, max_col, color="D0D7E3"):
    side = Side(style="thin", color=color)
    b = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b


def create_excel(data):
    wb = Workbook()

    # ======== 1. UMUMIY HISOBOT ========
    ws = wb.active
    ws.title = "Umumiy Hisobot"
    ws.sheet_view.showGridLines = False

    hdr(f"KUNLIK MOLIYAVIY HISOBOT", ws, 1, 1, 7, bg="0F2044", size=14, height=36)
    hdr(f"Sana: {TODAY}  |  Vaqt: {NOW}  |  {MONTH_NAME}", ws, 2, 1, 7, bg="1A3A6B", fg="BDD7EE", size=10, height=20)

    ws.row_dimensions[4].height = 18
    kpis = [
        ("JAMI BALANS", data["total_balance"], "16A34A"),
        ("BUGUNGI XARAJAT", data["total_today"], "DC2626"),
        ("OYLIK XARAJAT", data["total_month"], "D97706"),
        ("KOMPANIYALAR", data["companies_count"], "3B6EF6"),
        ("DIREKTORLAR", data["directors_count"], "7C3AED"),
        ("ISHCHILAR", data["workers_count"], "059669"),
        ("BUGUNGI TERIM", data["flowers_today_total"], "16A34A"),
    ]
    for i, (label, val, clr) in enumerate(kpis):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.cell(row=4, column=col, value=label).font = Font(bold=True, color=clr, size=9, name="Arial")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor="F0F4FF")
        v_cell = ws.cell(row=5, column=col, value=val)
        v_cell.font = Font(bold=True, color=clr, size=15, name="Arial")
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        v_cell.fill = PatternFill("solid", fgColor="F8FAFF")
        if i < 3:
            v_cell.number_format = '#,##0 "so\'m"'
        ws.row_dimensions[5].height = 30
    border_all(ws, 4, 5, 1, 7)

    # Kompaniyalar jadvali
    ws.row_dimensions[7].height = 14
    hdr("KOMPANIYALAR BALANSI VA XARAJATLAR", ws, 8, 1, 7, bg="0F2044", size=11, height=28)
    col_heads = ["#", "Kompaniya", "Direktor", "Joriy Balans", "Bugungi Xarajat", "Oylik Xarajat", "Bugun Terim"]
    col_widths = [5, 26, 18, 22, 22, 22, 18]
    for j, (h, w) in enumerate(zip(col_heads, col_widths), 1):
        cell = ws.cell(row=9, column=j, value=h)
        styled(cell, bold=True, color="FFFFFF", bg="1E3A5F", align="center", size=10)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[9].height = 20

    for i, comp in enumerate(data["companies"], 1):
        row = 9 + i
        bg = "F0F7FF" if i % 2 == 0 else "FFFFFF"
        # Bu kompaniyaning bugungi terimi
        comp_flowers = sum(
            fc["total"] for fc in data["flowers_today_by_company"]
            if fc["company"] == comp["name"]
        )
        vals = [i, comp["name"], comp["director"] or "—",
                comp["balance"], comp["today_exp"], comp["month_exp"], comp_flowers]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            align = "center" if j in [1, 3] else ("right" if j >= 4 else "left")
            styled(cell, bg=bg, align=align, size=10)
            if j >= 4:
                cell.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    last_data_row = 9 + len(data["companies"])
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=2, value="JAMI").font = Font(bold=True, size=10, name="Arial")
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="E8F4FD")
    for col_idx, key in [(4, "total_balance"), (5, "total_today"), (6, "total_month")]:
        cell = ws.cell(row=total_row, column=col_idx, value=data[key])
        cell.font = Font(bold=True, color="0F2044", size=11, name="Arial")
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = PatternFill("solid", fgColor="E8F4FD")
    flowers_total_cell = ws.cell(row=total_row, column=7, value=data["flowers_today_total"])
    flowers_total_cell.font = Font(bold=True, color="16A34A", size=11, name="Arial")
    flowers_total_cell.number_format = '#,##0'
    flowers_total_cell.alignment = Alignment(horizontal="right")
    flowers_total_cell.fill = PatternFill("solid", fgColor="E8F4FD")
    ws.row_dimensions[total_row].height = 22
    border_all(ws, 9, total_row, 1, 7)

    # Bugungi xarajatlar
    start = total_row + 2
    hdr("BUGUNGI XARAJATLAR", ws, start, 1, 7, bg="7F1D1D", fg="FFFFFF", size=11, height=28)
    exp_heads = ["#", "Vaqt", "Kompaniya", "Direktor", "Summa", "Izoh", ""]
    exp_widths = [5, 18, 24, 18, 20, 35, 5]
    for j, (h, w) in enumerate(zip(exp_heads, exp_widths), 1):
        cell = ws.cell(row=start+1, column=j, value=h)
        styled(cell, bold=True, color="FFFFFF", bg="991B1B", align="center", size=10)
        ws.column_dimensions[get_column_letter(j)].width = max(ws.column_dimensions[get_column_letter(j)].width, w)
    ws.row_dimensions[start+1].height = 20

    if data["today_expenses"]:
        for i, exp in enumerate(data["today_expenses"], 1):
            row = start + 1 + i
            bg = "FFF5F5" if i % 2 == 0 else "FFFFFF"
            vals = [i, str(exp["created_at"])[:16], exp["company"],
                    exp["director"] or "—", exp["amount"], exp["description"] or "—", ""]
            for j, v in enumerate(vals, 1):
                align = "center" if j in [1, 2, 4] else ("right" if j == 5 else "left")
                cell = ws.cell(row=row, column=j, value=v)
                styled(cell, bg=bg, align=align, size=9)
                if j == 5:
                    cell.number_format = '#,##0'
            ws.row_dimensions[row].height = 16

        exp_total_row = start + 1 + len(data["today_expenses"]) + 1
        ws.cell(row=exp_total_row, column=4, value="JAMI:").font = Font(bold=True, size=10, name="Arial")
        t_cell = ws.cell(row=exp_total_row, column=5, value=data["total_today"])
        t_cell.font = Font(bold=True, color="DC2626", size=11, name="Arial")
        t_cell.number_format = '#,##0'
        t_cell.alignment = Alignment(horizontal="right")
        for col_i in range(1, 8):
            ws.cell(row=exp_total_row, column=col_i).fill = PatternFill("solid", fgColor="FFE4E4")
        border_all(ws, start+1, exp_total_row, 1, 6)
    else:
        ws.cell(row=start+2, column=1, value="Bugun xarajat yo'q").font = Font(italic=True, color="999999", name="Arial")
        ws.merge_cells(f"A{start+2}:G{start+2}")

    # ======== 2. GULLAR TERIMI SHEET ========
    wg = wb.create_sheet("Gullar Terimi")
    wg.sheet_view.showGridLines = False

    hdr("GULLAR TERIMI HISOBOTI", wg, 1, 1, 5, bg="14532D", size=13, height=34)
    hdr(f"Sana: {TODAY}  |  {MONTH_NAME}", wg, 2, 1, 5, bg="166534", fg="BBF7D0", size=10, height=20)

    # KPI
    gul_kpis = [
        ("BUGUNGI TERIM", data["flowers_today_total"], "16A34A"),
        ("OYLIK TERIM", data["flowers_month_total"], "15803D"),
    ]
    for i, (label, val, clr) in enumerate(gul_kpis):
        col = i * 3 + 1
        end_col = col + 2
        wg.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(end_col)}4")
        wg.cell(row=4, column=col, value=label).font = Font(bold=True, color=clr, size=10, name="Arial")
        wg.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        wg.cell(row=4, column=col).fill = PatternFill("solid", fgColor="F0FDF4")
        wg.merge_cells(f"{get_column_letter(col)}5:{get_column_letter(end_col)}5")
        v = wg.cell(row=5, column=col, value=f"{val:,} ta")
        v.font = Font(bold=True, color=clr, size=18, name="Arial")
        v.alignment = Alignment(horizontal="center", vertical="center")
        v.fill = PatternFill("solid", fgColor="DCFCE7")
        wg.row_dimensions[5].height = 34
    for col in range(1, 6):
        wg.column_dimensions[get_column_letter(col)].width = 20
    border_all(wg, 4, 5, 1, 5)

    # Gul turlari jadvali
    wg.row_dimensions[7].height = 14
    hdr("GUL TURLARI BO'YICHA BUGUNGI TERIM", wg, 8, 1, 5, bg="14532D", size=11, height=26)
    g_heads = ["#", "Gul nomi", "1-sort", "2-sort", "Jami"]
    for j, h in enumerate(g_heads, 1):
        cell = wg.cell(row=9, column=j, value=h)
        styled(cell, bold=True, color="FFFFFF", bg="166534", align="center", size=10)
    wg.row_dimensions[9].height = 20

    if data["flowers_today_by_type"]:
        for i, f in enumerate(data["flowers_today_by_type"], 1):
            row = 9 + i
            bg = "F0FDF4" if i % 2 == 0 else "FFFFFF"
            vals = [i, f["flower_name"], f["s1"] or 0, f["s2"] or 0, f["total"]]
            for j, v in enumerate(vals, 1):
                align = "left" if j == 2 else "center"
                cell = wg.cell(row=row, column=j, value=v)
                styled(cell, bg=bg, align=align, size=10)
                if j >= 3:
                    cell.number_format = '#,##0'
            wg.row_dimensions[row].height = 18

        gt_row = 9 + len(data["flowers_today_by_type"]) + 1
        wg.cell(row=gt_row, column=2, value="JAMI").font = Font(bold=True, color="14532D", size=10, name="Arial")
        tc = wg.cell(row=gt_row, column=5, value=data["flowers_today_total"])
        tc.font = Font(bold=True, color="14532D", size=12, name="Arial")
        tc.number_format = '#,##0'
        tc.alignment = Alignment(horizontal="center")
        for ci in range(1, 6):
            wg.cell(row=gt_row, column=ci).fill = PatternFill("solid", fgColor="DCFCE7")
        border_all(wg, 9, gt_row, 1, 5)
    else:
        wg.cell(row=10, column=1, value="Bugun terim kiritilmagan").font = Font(italic=True, color="999999", name="Arial")
        wg.merge_cells("A10:E10")

    # Kompaniya bo'yicha terim
    if data["flowers_today_by_company"]:
        fc_start = 9 + len(data["flowers_today_by_type"]) + 4
        hdr("KOMPANIYA BO'YICHA TERIM", wg, fc_start, 1, 5, bg="14532D", size=11, height=26)
        fc_heads = ["#", "Kompaniya", "", "", "Jami"]
        for j, h in enumerate(fc_heads, 1):
            cell = wg.cell(row=fc_start+1, column=j, value=h)
            styled(cell, bold=True, color="FFFFFF", bg="166534", align="center", size=10)
        for i, fc in enumerate(data["flowers_today_by_company"], 1):
            row = fc_start + 1 + i
            bg = "F0FDF4" if i % 2 == 0 else "FFFFFF"
            wg.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
            wg.cell(row=row, column=2, value=fc["company"])
            tc = wg.cell(row=row, column=5, value=fc["total"])
            tc.number_format = '#,##0'
            tc.alignment = Alignment(horizontal="center")
            for ci in range(1, 6):
                wg.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=bg)
            wg.row_dimensions[row].height = 18
        border_all(wg, fc_start, fc_start + 1 + len(data["flowers_today_by_company"]), 1, 5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def create_pdf(data):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    def h1(text): return Paragraph(text, ParagraphStyle('h1', parent=styles['Normal'],
        fontSize=16, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f2044'), spaceAfter=4))
    def h2(text): return Paragraph(text, ParagraphStyle('h2', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a5f'), spaceAfter=6, spaceBefore=10))
    def sub(text): return Paragraph(text, ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=10))

    story = []

    # Sarlavha
    story.append(h1("KUNLIK MOLIYAVIY HISOBOT"))
    story.append(sub(f"Sana: {TODAY}   |   Vaqt: {NOW}   |   {MONTH_NAME}"))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#3b6ef6')))
    story.append(Spacer(1, 8))

    # KPI
    story.append(h2("Umumiy Ko'rsatkichlar"))
    kpi_data = [
        ["Ko'rsatkich", "Qiymat", "Ko'rsatkich", "Qiymat"],
        ["Jami balans", f"{data['total_balance']:,} so'm", "Kompaniyalar", str(data['companies_count'])],
        ["Bugungi xarajat", f"{data['total_today']:,} so'm", "Direktorlar", str(data['directors_count'])],
        ["Oylik xarajat", f"{data['total_month']:,} so'm", "Faol ishchilar", str(data['workers_count'])],
        ["Bugungi terim", f"{data['flowers_today_total']:,} ta", "Oylik terim", f"{data['flowers_month_total']:,} ta"],
    ]
    kt = Table(kpi_data, colWidths=[5*cm, 5*cm, 5*cm, 3*cm])
    kt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f2044')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f0f7ff'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#dee4f0')),
        ('PADDING', (0,0), (-1,-1), 7),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('FONTNAME', (1,1), (1,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1,1), (1,1), colors.HexColor('#16a34a')),
        ('TEXTCOLOR', (1,2), (1,2), colors.HexColor('#dc2626')),
        ('TEXTCOLOR', (1,3), (1,3), colors.HexColor('#d97706')),
        ('TEXTCOLOR', (1,4), (1,4), colors.HexColor('#16a34a')),
    ]))
    story.append(kt)
    story.append(Spacer(1, 10))

    # Gullar terimi
    if data["flowers_today_by_type"]:
        story.append(h2("Bugungi Gullar Terimi"))
        f_rows = [["#", "Gul nomi", "1-sort", "2-sort", "Jami"]]
        for i, f in enumerate(data["flowers_today_by_type"], 1):
            f_rows.append([str(i), f["flower_name"],
                           f"{f['s1']:,}", f"{f['s2']:,}", f"{f['total']:,}"])
        f_rows.append(["", "JAMI", "", "", f"{data['flowers_today_total']:,}"])
        ft = Table(f_rows, colWidths=[1*cm, 5*cm, 3*cm, 3*cm, 3*cm])
        ft.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14532d')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor('#f0fdf4'), colors.white]),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#dcfce7')),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (4,-1), (4,-1), colors.HexColor('#16a34a')),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#bbf7d0')),
            ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('PADDING', (0,0), (-1,-1), 6),
            ('FONTNAME', (0,1), (-1,-2), 'Helvetica'),
        ]))
        story.append(ft)
        story.append(Spacer(1, 10))

    # Kompaniyalar
    story.append(h2("Kompaniyalar Balansi"))
    comp_rows = [["#", "Kompaniya", "Direktor", "Balans", "Bugun xarajat", "Bugun terim"]]
    for i, comp in enumerate(data["companies"], 1):
        comp_flowers = sum(
            fc["total"] for fc in data["flowers_today_by_company"]
            if fc["company"] == comp["name"]
        )
        comp_rows.append([
            str(i), comp["name"], comp["director"] or "—",
            f"{comp['balance']:,}", f"{comp['today_exp']:,}", f"{comp_flowers:,}"
        ])
    comp_rows.append(["", "JAMI", "",
                       f"{data['total_balance']:,}",
                       f"{data['total_today']:,}",
                       f"{data['flowers_today_total']:,}"])

    ct = Table(comp_rows, colWidths=[0.7*cm, 4.5*cm, 3*cm, 3.2*cm, 3*cm, 3.6*cm])
    ct.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f2044')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor('#f0f7ff'), colors.white]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e0ecff')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#dee4f0')),
        ('ALIGN', (3,0), (5,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('FONTNAME', (0,1), (2,-1), 'Helvetica'),
    ]))
    story.append(ct)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def main():
    global TODAY, NOW, MONTH, MONTH_NAME
    TODAY = datetime.now().strftime("%Y-%m-%d")
    NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
    MONTH = datetime.now().strftime("%Y-%m")
    MONTH_NAME = datetime.now().strftime("%B %Y")

    print(f"[{NOW}] Kunlik hisobot yuborilmoqda...")
    data = get_report_data()
    send_daily_telegram(data)
    excel_bytes = create_excel(data)
    send_telegram_file(excel_bytes, f"hisobot_{TODAY}.xlsx", f"📊 Excel hisobot — {TODAY}")
    pdf_bytes = create_pdf(data)
    send_telegram_file(pdf_bytes, f"hisobot_{TODAY}.pdf", f"📄 PDF hisobot — {TODAY}")
    print(f"[{NOW}] Hisobot muvaffaqiyatli yuborildi!")


if __name__ == "__main__":
    main()